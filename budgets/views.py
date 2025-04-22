import io
import json
from calendar import calendar
from collections import defaultdict
from datetime import datetime, time
from decimal import Decimal
from functools import reduce
from operator import or_
from time import sleep

import openpyxl
import pandas as pd
import xlsxwriter
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import check_password

from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Count, Sum, F, Q, Subquery, OuterRef
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from openpyxl.styles import colors, Font
from openpyxl.utils.cell import get_column_letter
from openpyxl.workbook import Workbook
from pandas import DataFrame
from rest_framework_simplejwt.tokens import RefreshToken

from BudgetForecaster import settings
from budgets.forms import BudgetEditForm, UserCreate
from budgets.models import BudgetLines, Users, Glafs, Glpjd, BudgetComments, BudgetStatus, BudgetAssumptions, \
    BudgetTotals, \
    Currency, BudgetLinesLog, Glamf, BudgetVariations, Department, Enebd, Enrqnl, Poporl, Accounts, Poporh1, Enebh, \
    ChangeLog, FinancialYear, Group, BudgetLinesLogVariations


def netperd_in_quarter(quarter):
    quarter = int(quarter)
    if quarter == 1:
        return ['netperd1', 'netperd2', 'netperd3']
    elif quarter == 2:
        return ['netperd4', 'netperd5', 'netperd6']
    elif quarter == 3:
        return ['netperd7', 'netperd8', 'netperd9']
    elif quarter == 4:
        return ['netperd10', 'netperd11', 'netperd12']
    else:
        return []


def netperd_in_half(quarter):
    quarter = int(quarter)
    if quarter == 1:
        return ['netperd1', 'netperd2', 'netperd3', 'netperd4', 'netperd5', 'netperd6']
    elif quarter == 2:
        return ['netperd7', 'netperd8', 'netperd9', 'netperd10', 'netperd11', 'netperd12']
    else:
        return []


def months_in_quarter(quarter):
    quarter = int(quarter)
    if quarter == 1:
        return ['January', 'February', 'March']
    elif quarter == 2:
        return ['April', 'May', 'June']
    elif quarter == 3:
        return ['July', 'August', 'September']
    elif quarter == 4:
        return ['October', 'November', 'December']
    else:
        return []


def current_quarter():
    now = datetime.now()
    month = now.month

    if month in range(1, 4):
        quarter = 1
    elif month in range(4, 7):
        quarter = 2
    elif month in range(7, 10):
        quarter = 3
    else:
        quarter = 4

    return quarter


def current_period():
    now = datetime.now()
    month = now.month
    period_dict = {
        1: 'netperd1',
        2: 'netperd2',
        3: 'netperd3',
        4: 'netperd4',
        5: 'netperd5',
        6: 'netperd6',
        7: 'netperd7',
        8: 'netperd8',
        9: 'netperd9',
        10: 'netperd10',
        11: 'netperd11',
        12: 'netperd12'
    }
    return period_dict.get(month, None)


def user_login(request):
    if request.method == 'POST':
        try:
            username = request.POST["username"]
            password = request.POST["password"]
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                if user.role != '002':
                    # active = BudgetStatus.objects.get(department=request.user.department, is_active=True)
                    return redirect('budgets:dashboard-home-department')
                """user_obj = get_object_or_404(Users, username=username)
                    totp = pyotp.TOTP(user.otp_base32).now()
                    user_obj.login_otp = totp
                    user_obj.otp_created_at = datetime.now(timezone.utc)
                    user_obj.login_otp_used = False
                    user_obj.save(update_fields=["login_otp", "otp_created_at", "login_otp_used"])
                    #return redirect('budgets:enter-otp')"""
                return redirect('budgets:dashboard-home')
            else:
                messages.error(request, message="Username/Password not registered")
                return redirect("budgets:login")

        except Exception as e:
            print(e)
            messages.error(request, message="An error occurred. Please try again.")

            return redirect("budgets:login")

    else:

        return render(request, 'login.html')


# Create your views here.
def custom_404(request, exception):
    return render(request, '404.html', status=404)


def custom_500(request, exception):
    return render(request, '500.html', status=500)


def get_actual_amounts(obj, current_q):
    """
    Takes account ids and queries the GLPJD table for details of posted transactions and returns a list of values
    :param obj:
    :param netperd:
    :return: actual_data
    """
    actual_data = []
    year = FinancialYear.objects.get(is_active=True)

    if 'netperd' in current_q:
        ranges = {
            'netperd1': f'audtdate__gte={year.year}0101,audtdate__lte={year.year}0131',
            'netperd2': f'audtdate__gte={year.year}0201,audtdate__lte={year.year}0229',
            'netperd3': f'audtdate__gte={year.year}0301,audtdate__lte={year.year}0331',
            'netperd4': f'audtdate__gte={year.year}0401,audtdate__lte={year.year}0430',
            'netperd5': f'audtdate__gte={year.year}0501,audtdate__lte={year.year}0531',
            'netperd6': f'audtdate__gte={year.year}0601,audtdate__lte={year.year}0630',
            'netperd7': f'audtdate__gte={year.year}0701,audtdate__lte={year.year}0731',
            'netperd8': f'audtdate__gte={year.year}0801,audtdate__lte={year.year}0831',
            'netperd9': f'audtdate__gte={year.year}0901,audtdate__lte={year.year}0930',
            'netperd10': f'audtdate__gte={year.year}1001,audtdate__lte={year.year}1031',
            'netperd11': f'audtdate__gte={year.year}1101,audtdate__lte={year.year}1130',
            'netperd12': f'audtdate__gte={year.year}1201,audtdate__lte={year.year}1231',
        }
        if obj:
            start_date, end_date = ranges[current_q].split(',')
            start_date = f'{year.year}0101'
            end_date = int(end_date.split('=')[1])
            acctids = obj

            actual_transactions = Glpjd.objects.using('sql_server').filter(
                acctid__icontains=acctids,
                audtdate__gte=start_date,
                audtdate__lte=end_date
            ).exclude(transamt=0).values('jnldtldesc', 'jnldtlref', 'acctid', 'audtdate', 'transamt')

            for expense in actual_transactions:
                date_obj = datetime.strptime(str(expense['audtdate']), '%Y%m%d')
                formatted_date = date_obj.strftime('%d-%m-%Y')

                if expense is not None:  # Check if document number is not empty
                    actual_data.append({
                        'amount': expense['transamt'],
                        'description': expense['jnldtldesc'],
                        'account_id': expense['acctid'],
                        'date': formatted_date,
                        'document_number': expense['jnldtlref']
                    })

        return actual_data

    else:
        ranges = {
            '1': f'audtdate__gte={year.year}0101,audtdate__lte={year.year}0331',

            '2': f'audtdate__gte={year.year}0401,audtdate__lte={year.year}0630',

            '3': f'audtdate__gte={year.year}0701,audtdate__lte={year.year}0930',

            '4': f'audtdate__gte={year.year}1001,audtdate__lte={year.year}1231',

        }
        if obj:
            start_date, end_date = ranges[current_q].split(',')
            start_date = f'{year.year}0101'
            end_date = int(end_date.split('=')[1])
            acctids = obj

            actual_transactions = Glpjd.objects.using('sql_server').filter(
                acctid__icontains=acctids,
                audtdate__gte=start_date,
                audtdate__lte=end_date
            ).exclude(transamt=0).values('jnldtldesc', 'jnldtlref', 'acctid', 'audtdate', 'transamt')

            for expense in actual_transactions:
                date_obj = datetime.strptime(str(expense['audtdate']), '%Y%m%d')
                formatted_date = date_obj.strftime('%d-%m-%Y')

                if expense is not None:  # Check if document number is not empty
                    actual_data.append({
                        'amount': expense['transamt'],
                        'description': expense['jnldtldesc'],
                        'account_id': expense['acctid'],
                        'date': formatted_date,
                        'document_number': expense['jnldtlref']
                    })

        return actual_data


def get_pending_amounts(obj, current_q):
    """
    Takes account ids and queries sql server database for expense vouchers, purchase orders and purchase requisitions that are pending and returns a list of them
    :param obj:
    :param netperd:
    :return: pending_data
    """
    year = FinancialYear.objects.get(is_active=True)

    if 'netperd' in current_q:
        ranges = {
            'netperd1': f'audtdate__gte={year.year}0101,audtdate__lte={year.year}0131',
            'netperd2': f'audtdate__gte={year.year}0201,audtdate__lte={year.year}0229',
            'netperd3': f'audtdate__gte={year.year}0301,audtdate__lte={year.year}0331',
            'netperd4': f'audtdate__gte={year.year}0401,audtdate__lte={year.year}0430',
            'netperd5': f'audtdate__gte={year.year}0501,audtdate__lte={year.year}0531',
            'netperd6': f'audtdate__gte={year.year}0601,audtdate__lte={year.year}0630',
            'netperd7': f'audtdate__gte={year.year}0701,audtdate__lte={year.year}0731',
            'netperd8': f'audtdate__gte={year.year}0801,audtdate__lte={year.year}0831',
            'netperd9': f'audtdate__gte={year.year}0901,audtdate__lte={year.year}0930',
            'netperd10': f'audtdate__gte={year.year}1001,audtdate__lte={year.year}1031',
            'netperd11': f'audtdate__gte={year.year}1101,audtdate__lte={year.year}1130',
            'netperd12': f'audtdate__gte={year.year}1201,audtdate__lte={year.year}1231',
        }
        pending_data = []
        if obj:
            start_date, end_date = ranges[current_q].split(',')
            start_date = int(year.year + '0101')
            end_date = int(end_date.split('=')[1])
            acctids = obj

            pending_expenses = Enebd.objects.using('sql_server').filter(
                idglacct__icontains=acctids,
                status=1,
                expdate__gte=start_date,
                expdate__lte=end_date,

            ).exclude(amtlinet=0).values('amtlinet', 'textdesc', 'idglacct', 'audtdate', 'cntbtch')
            pending_expenses_numbers = Enebh.objects.using('sql_server').filter(
                cntbtch__in=pending_expenses.values_list('cntbtch', flat=True)
            ).values('idexpst', 'cntbtch')
            for expense in pending_expenses:
                date_obj = datetime.strptime(str(expense['audtdate']), '%Y%m%d')
                formatted_date = date_obj.strftime('%d-%m-%Y')
                document_number = next((number['idexpst'] for number in pending_expenses_numbers if
                                        number['cntbtch'] == expense['cntbtch']), None)
                if document_number is not None:  # Check if document number is not empty
                    pending_data.append({
                        'transaction_type': 'Expense Voucher',
                        'amount': expense['amtlinet'],
                        'description': expense['textdesc'],
                        'account_id': expense['idglacct'],
                        'date': formatted_date,
                        'document_number': document_number
                    })
            pending_requisition =	Enrqnl.objects.using('sql_server').filter(
                fmtglacct__icontains=acctids,
                status=1,
                exparrival__gte=start_date,
                exparrival__lte=end_date,

            ).exclude(extended = 0).values('itemdesc', 'extended', 'fmtglacct', 'audtdate', 'rqnnumber','status')
            print(pending_requisition)
            for req in pending_requisition:
                date_obj = datetime.strptime(str(req['audtdate']), '%Y%m%d')
                formatted_date = date_obj.strftime('%d-%m-%Y')
                document_number = req['rqnnumber']
                if document_number is not None:  # Check if document number is not empty
                    pending_data.append({
                        'transaction_type': 'Purchase Requisition',
                        'amount': req['extended'],
                        'description': req['itemdesc'],
                        'account_id': req['fmtglacct'],
                        'date': formatted_date,
                        'document_number': document_number
                    })

            pending_purchase_order = Poporl.objects.using('sql_server').filter(
                glnonstkcr__icontains=acctids,
                exparrival__gte=start_date,
                exparrival__lte=end_date,
                completion=1

            ).exclude(extended =0).values('itemdesc', 'extended', 'glnonstkcr', 'audtdate', 'porhseq')
            pending_purchase_order_header = Poporh1.objects.using('sql_server').filter(
                porhseq__in=pending_purchase_order.values_list('porhseq', flat=True)
            ).values('ponumber', 'porhseq')
            po_number_mapping = {header['porhseq']: header['ponumber'] for header in pending_purchase_order_header}
            for po in pending_purchase_order:
                date_obj = datetime.strptime(str(po['audtdate']), '%Y%m%d')
                formatted_date = date_obj.strftime('%d-%m-%Y')
                document_number = po_number_mapping.get(po['porhseq'])
                if document_number is not None:  # Check if document number is not empty
                    pending_data.append({
                        'transaction_type': 'Purchase Order',
                        'amount': po['extended'],
                        'description': po['itemdesc'],
                        'account_id': po['glnonstkcr'],
                        'date': formatted_date,
                        'document_number': document_number
                    })

        return pending_data

    else:
        ranges = {
            '1': f'audtdate__gte={year.year}0101,audtdate__lte={year.year}0331',

            '2': f'audtdate__gte={year.year}0401,audtdate__lte={year.year}0630',

            '3': f'audtdate__gte={year.year}0701,audtdate__lte={year.year}0930',

            '4': f'audtdate__gte={year.year}1001,audtdate__lte={year.year}1231',

        }
        pending_data = []
        if obj:
            start_date, end_date = ranges[current_q].split(',')
            start_date = int(year.year + '0101')
            end_date = int(end_date.split('=')[1])
            acctids = obj

            pending_expenses = Enebd.objects.using('sql_server').filter(
                idglacct__icontains=acctids,
                status=1,
                expdate__gte=start_date,
                expdate__lte=end_date,

            ).exclude(amtlinet=0).values('amtlinet', 'textdesc', 'idglacct', 'audtdate', 'cntbtch')
            pending_expenses_numbers = Enebh.objects.using('sql_server').filter(
                cntbtch__in=pending_expenses.values_list('cntbtch', flat=True)
            ).values('idexpst', 'cntbtch')

            for expense in pending_expenses:
                date_obj = datetime.strptime(str(expense['audtdate']), '%Y%m%d')
                formatted_date = date_obj.strftime('%d-%m-%Y')
                document_number = next((number['idexpst'] for number in pending_expenses_numbers if
                                        number['cntbtch'] == expense['cntbtch']), None)
                if document_number is not None:  # Check if document number is not empty
                    pending_data.append({
                        'transaction_type': 'Expense Voucher',
                        'amount': expense['amtlinet'],
                        'description': expense['textdesc'],
                        'account_id': expense['idglacct'],
                        'date': formatted_date,
                        'document_number': document_number
                    })
            pending_requisition = Enrqnl.objects.using('sql_server').filter(
                fmtglacct__icontains=acctids,
                status=1,
                exparrival__gte=start_date,
                exparrival__lte=end_date,

            ).exclude(extended =0).values('itemdesc', 'extended', 'fmtglacct', 'audtdate', 'rqnnumber')

            for req in pending_requisition:
                date_obj = datetime.strptime(str(req['audtdate']), '%Y%m%d')
                formatted_date = date_obj.strftime('%d-%m-%Y')
                document_number = req['rqnnumber']
                if document_number is not None:  # Check if document number is not empty
                    pending_data.append({
                        'transaction_type': 'Purchase Requisition',
                        'amount': req['extended'],
                        'description': req['itemdesc'],
                        'account_id': req['fmtglacct'],
                        'date': formatted_date,
                        'document_number': document_number
                    })

            pending_purchase_order = Poporl.objects.using('sql_server').filter(
                glnonstkcr__icontains=acctids,
                completion=1,
                exparrival__gte=start_date,
                exparrival__lte=end_date,

            ).exclude(extended=0).values('itemdesc', 'extended', 'glnonstkcr', 'audtdate', 'porhseq')
            pending_purchase_order_header = Poporh1.objects.using('sql_server').filter(
                porhseq__in=pending_purchase_order.values_list('porhseq', flat=True)
            ).values('ponumber', 'porhseq')
            po_number_mapping = {header['porhseq']: header['ponumber'] for header in pending_purchase_order_header}
            for po in pending_purchase_order:
                date_obj = datetime.strptime(str(po['audtdate']), '%Y%m%d')
                formatted_date = date_obj.strftime('%d-%m-%Y')
                document_number = po_number_mapping.get(po['porhseq'])
                if document_number is not None:  # Check if document number is not empty
                    pending_data.append({
                        'transaction_type': 'Purchase Order',
                        'amount': po['extended'],
                        'description': po['itemdesc'],
                        'account_id': po['glnonstkcr'],
                        'date': formatted_date,
                        'document_number': document_number
                    })

        return pending_data

def format_value(value):
    """Helper function to format numbers to comma-separated string with two decimals."""
    try:
        return "{:,.2f}".format(Decimal(value))
    except (TypeError, ValueError):
        return value
def period_aggregate_singular(obj, obj2, netperd):
    """

    :param obj: object of the budget total queryset
    :param obj2: object of the actual queryset
    :param netperd: the period from which our values should come
    :return period_values_dict
    """
    year = FinancialYear.objects.get(is_active=True)

    period_values_dict = {
        'total': 0,
        'total_actual': 0,
        'available': 0,
        'pending': 0,
        'year_total': 0,
        'year_total_actual': 0,
        'pending_year_total': 0,
        'available_year_total': 0
    }
    ranges = {
        'netperd1': f'audtdate__gte={year.year}0101,audtdate__lte={year.year}0131',
        'netperd2': f'audtdate__gte={year.year}0201,audtdate__lte={year.year}0229',
        'netperd3': f'audtdate__gte={year.year}0301,audtdate__lte={year.year}0331',
        'netperd4': f'audtdate__gte={year.year}0401,audtdate__lte={year.year}0430',
        'netperd5': f'audtdate__gte={year.year}0501,audtdate__lte={year.year}0531',
        'netperd6': f'audtdate__gte={year.year}0601,audtdate__lte={year.year}0630',
        'netperd7': f'audtdate__gte={year.year}0701,audtdate__lte={year.year}0731',
        'netperd8': f'audtdate__gte={year.year}0801,audtdate__lte={year.year}0831',
        'netperd9': f'audtdate__gte={year.year}0901,audtdate__lte={year.year}0930',
        'netperd10': f'audtdate__gte={year.year}1001,audtdate__lte={year.year}1031',
        'netperd11': f'audtdate__gte={year.year}1101,audtdate__lte={year.year}1130',
        'netperd12': f'audtdate__gte={year.year}1201,audtdate__lte={year.year}1231',
    }

    if obj and obj2:
        # Calculating quarter total
        period_total = obj.get(netperd, 0)
        # Calculating year total
        year_total = sum(getattr(obj2, f"netperd{i}", 0) for i in range(1, int(netperd.replace('netperd', '')) + 1))
        period_values_dict['total'] = period_total
        period_values_dict['year_total'] = year_total
        start_date, end_date = ranges[netperd].split(',')
        start_date = int(start_date.split('=')[1])
        end_date = int(end_date.split('=')[1])
        acctids = obj['account__acctfmttd']

        pending_expenses = Enebd.objects.using('sql_server').filter(
            idglacct__icontains=acctids,
            status=1,
            expdate__gte=start_date,
            expdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_expenses_all = Enebd.objects.using('sql_server').filter(
            idglacct__icontains=acctids,
            status=1,
            expdate__gte=int(year.year + '0101'),
            expdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_requisition = Enrqnl.objects.using('sql_server').filter(
            fmtglacct__icontains=acctids,
            status=1,
            exparrival__gte=start_date,
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_requisition_all = Enrqnl.objects.using('sql_server').filter(
            fmtglacct__icontains=acctids,
            status=1,
            exparrival__gte=int(year.year + '0101'),
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order = Poporl.objects.using('sql_server').filter(
            glnonstkcr__icontains=acctids,
            completion=1,
            exparrival__gte=start_date,
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0
        pending_purchase_order_all = Poporl.objects.using('sql_server').filter(
            glnonstkcr__icontains=acctids,
            completion=1,
            exparrival__gte=int(year.year + '0101'),
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_total = pending_expenses + pending_requisition + pending_purchase_order
        print(pending_total)
        pending_year_total = pending_expenses_all + pending_requisition_all + pending_purchase_order_all
        available_year = round(year_total, 2) - round(pending_year_total, 2)
        available = round(period_total, 2) - round(pending_total, 2)
        period_values_dict['pending'] = pending_total
        period_values_dict['pending_year_total'] = pending_year_total
        period_values_dict['available'] = available
        period_values_dict['available_year_total'] = available_year
        period_values_dict['total'] = period_total
        period_values_dict['year_total'] = year_total
        acctidspec = obj['account_id']
        budget = Glafs.objects.using('sql_server').filter(acctid__icontains=acctidspec, fscsdsg='A', fscscurn='ZMW',
                                                          curntype='F',
                                                          fscsyr=year.year).all()
        budget_specific = budget.values(netperd, 'acctid')
        for obj, obj2 in zip(budget_specific, budget):
            if obj and obj2:
                period_total = obj.get(netperd, 0)
                # Calculating year total
                year_total = sum(
                    getattr(obj2, f"netperd{i}", 0) for i in range(1, int(netperd.replace('netperd', '')) + 1))

                period_values_dict['total_actual'] = period_total
                period_values_dict['year_total_actual'] = year_total
        total_actual = Decimal(period_values_dict['total_actual'])
        year_total_actual = Decimal(period_values_dict['year_total_actual'])
        pending_total = Decimal(period_values_dict['pending'])
        year_pending_total = Decimal(period_values_dict['pending_year_total'])

        available = round(Decimal(period_values_dict['total']), 2) - round(total_actual, 2) - round(
            pending_total, 2)
        year_available = round(Decimal(period_values_dict['year_total']), 2) - round(year_total_actual,
                                                                                                      2) - round(
            year_pending_total, 2)
        period_values_dict['total_actual']
        period_values_dict['available'] = available
        period_values_dict['available_year_total'] = year_available
    period_values_dict = {k: format_value(v) for k, v in period_values_dict.items()}
    return period_values_dict


def quarter_aggregate_singular(obj, obj2, current_q):
    """
    Similar to period_aggregate_singular but provides information according to user selected quarter
    :param obj:
    :param obj2:
    :param current_q:
    :return:
    """
    year = FinancialYear.objects.get(is_active=True)

    current_q = int(current_q)
    period_values_dict = {
        'total': Decimal('0.00'),
        'total_actual': Decimal('0.00'),
        'available': Decimal('0.00'),
        'pending': Decimal('0.00'),
        'year_total': Decimal('0.00'),
        'year_total_actual': Decimal('0.00'),
        'pending_year_total': Decimal('0.00'),
        'available_year_total': Decimal('0.00')
    }
    ranges = {
        1: f'audtdate__gte={year.year}0101,audtdate__lte={year.year}0331',

        2: f'audtdate__gte={year.year}0401,audtdate__lte={year.year}0630',

        3: f'audtdate__gte={year.year}0701,audtdate__lte={year.year}0930',

        4: f'audtdate__gte={year.year}1001,audtdate__lte={year.year}1231',

    }
    ytd_ranges = {
        1: 4,
        2: 7,
        3: 10,
        4: 13
    }
    months_in_current_quarter =netperd_in_quarter(current_q)
    net_perds = [month for month in months_in_current_quarter]
    period_values_aggregations = {period: Sum(period) for period in net_perds}
    if obj and obj2:
        period_values_sum = obj.aggregate(**period_values_aggregations)
        # Calculating year total
        netperd_fields = [f"netperd{i}" for i in range(1, ytd_ranges[current_q])]
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = obj2.aggregate(**total_period_values_aggregations)
        quarter_total = sum(period_values_sum.values())
        year_total = sum(total_period_values_sum.values())
        period_values_dict['quarter_total'] = quarter_total
        period_values_dict['year_total'] = year_total
        start_date, end_date = ranges[int(current_q)].split(',')
        start_date = int(start_date.split('=')[1])
        end_date = int(end_date.split('=')[1])
        acctids = [o['account__acctfmttd'] for o in obj]
        pending_expenses = Enebd.objects.using('sql_server').filter(
            idglacct__icontains__in=acctids,
            status=1,
            expdate__gte=start_date,
            expdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_expenses_all = Enebd.objects.using('sql_server').filter(
            idglacct__icontains__in=acctids,
            status=1,
            expdate__gte=int(year.year + '0101'),
            expdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_requisition = Enrqnl.objects.using('sql_server').filter(
            glacct__icontains__in=acctids,
            status=1,
            exparrival__gte=start_date,
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_requisition_all = Enrqnl.objects.using('sql_server').filter(
            glacct__icontains__in=acctids,
            status=1,
            exparrival__gte=int(year.year + '0101'),
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order = Poporl.objects.using('sql_server').filter(
            glnonstkcr__icontains__in=acctids,
            completion=1,
            exparrival__gte=start_date,
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order_all = Poporl.objects.using('sql_server').filter(
            glnonstkcr__icontains__in=acctids,
            completion=1,
            exparrival__gte=int(year.year + '0101'),
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_total = pending_expenses + pending_requisition + pending_purchase_order
        pending_year_total = pending_expenses_all + pending_requisition_all + pending_purchase_order_all
        available_year = round(year_total, 2) - round(pending_year_total, 2)
        available = round(quarter_total, 2) - round(pending_total, 2)
        period_values_dict['pending'] = "{:,.2f}".format(pending_total)
        period_values_dict['pending_year_total'] = "{:,.2f}".format(pending_year_total)
        period_values_dict['available'] = "{:,.2f}".format(available)
        period_values_dict['available_year_total'] = "{:,.2f}".format(available_year)
        period_values_dict['total'] = "{:,.2f}".format(quarter_total)
        period_values_dict['year_total'] = "{:,.2f}".format(year_total)
        budget = Glafs.objects.using('sql_server').filter(acctid=acctids, fscsdsg='A', fscscurn='ZMW', curntype='F',
                                                          fscsyr=year.year).all()
        budget_specific = budget.values(*net_perds, 'acctid')
        for obj, obj2 in zip(budget_specific, budget):
            if obj and obj2:
                quarter_total = sum(obj.get(period, 0) for period in net_perds)
                # Calculating year total
                year_total = sum(getattr(obj2, f"netperd{i}", 0) for i in range(1, ytd_ranges[current_q]))

                period_values_dict['total_actual'] = "{:,.2f}".format(quarter_total)
                period_values_dict['year_total_actual'] = "{:,.2f}".format(year_total)
        total_actual = Decimal(period_values_dict['total_actual'].replace(',', ''))
        year_total_actual = Decimal(period_values_dict['year_total_actual'].replace(',', ''))
        pending_total = Decimal(period_values_dict['pending'].replace(',', ''))
        year_pending_total = Decimal(period_values_dict['pending_year_total'].replace(',', ''))

        available = round(Decimal(period_values_dict['total'].replace(',', '')), 2) - round(total_actual, 2) - round(
            pending_total, 2)
        year_available = round(Decimal(period_values_dict['year_total'].replace(',', '')), 2) - round(
            year_total_actual,
            2) - round(
            year_pending_total, 2)
        period_values_dict['available'] = "{:,.2f}".format(available)
        period_values_dict['available_year_total'] = "{:,.2f}".format(year_available)

    return period_values_dict


def quarter_aggregate(queryset, queryset2, queryset3, queryset4, current_q, dept_id, ):
    """

    :param queryset: Budget totals quesryset for specific quarter
    :param queryset2: Budget totals queryset for year to date
    :param queryset3: Glafs queryset for specific quarter
    :param queryset4: Glafs queryset for yeart to date
    :param current_q: selected quarter
    :param dept_id:
    :return: period_values_dict
    """
    year = FinancialYear.objects.get(is_active=True)
    period_values_dict = {
        'total': Decimal('0.00'),
        'total_actual': Decimal('0.00'),
        'available': Decimal('0.00'),
        'pending': Decimal('0.00'),
        'year_total': Decimal('0.00'),
        'year_total_actual': Decimal('0.00'),
        'pending_year_total': Decimal('0.00'),
        'available_year_total': Decimal('0.00')
    }
    ytd_range = {
        1: 4,
        2: 7,
        3: 10,
        4: 13
    }
    ranges = {
        1: f'audtdate__gte={year.year}0101,audtdate__lte={year.year}0331',

        2: f'audtdate__gte={year.year}0401,audtdate__lte={year.year}0630',

        3: f'audtdate__gte={year.year}0701,audtdate__lte={year.year}0930',

        4: f'audtdate__gte={year.year}1001,audtdate__lte={year.year}1231',

    }

    months_in_current_quarter = netperd_in_quarter(current_q)
    net_perds = [month for month in months_in_current_quarter]
    period_values_aggregations = {period: Sum(period) for period in net_perds}
    if queryset and queryset2:
        period_values_sum = queryset.aggregate(**period_values_aggregations)
        netperd_fields = [f"netperd{i}" for i in range(1, ytd_range[int(current_q)])]
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = queryset2.aggregate(**total_period_values_aggregations)
        quarter_total = sum(period_values_sum.values())
        year_total = sum(total_period_values_sum.values())
        start_date, end_date = ranges[int(current_q)].split(',')
        start_date = int(start_date.split('=')[1])
        end_date = int(end_date.split('=')[1])
        acctids = [obj['account__acctfmttd'] for obj in queryset]
        q1 = reduce(or_, [Q(idglacct__icontains=item) for item in acctids])
        q2 = reduce(or_, [Q(glacct__icontains=item) for item in acctids])
        q3 = reduce(or_, [Q(glnonstkcr__icontains=item) for item in acctids])

        pending_expenses = Enebd.objects.using('sql_server').filter(
            q1,
            status=1,
            expdate__gte=start_date,
            expdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_expenses_all = Enebd.objects.using('sql_server').filter(
            q1,
            status=1,
            expdate__gte=int(year.year + '0101'),
            expdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_requisition = Enrqnl.objects.using('sql_server').filter(
            q2,
            status=1,
            exparrival__gte=start_date,
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_requisition_all = Enrqnl.objects.using('sql_server').filter(
            q2,
            status=1,
            exparrival__gte=int(year.year + '0101'),
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order = Poporl.objects.using('sql_server').filter(
            q3,
            completion=1,
            exparrival__gte=start_date,
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order_all = Poporl.objects.using('sql_server').filter(
            q3,
            completion=1,
            exparrival__gte=int(year.year + '0101'),
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_total = pending_expenses + pending_requisition + pending_purchase_order
        pending_year_total = pending_expenses_all + pending_requisition_all + pending_purchase_order_all
        period_values_dict['pending'] = round(pending_total, 2)
        period_values_dict['pending_year_total'] = round(pending_year_total, 2)
        period_values_dict['available'] = round(quarter_total, 2) - round(pending_total, 2)
        period_values_dict['available_year_total'] = round(year_total, 2) - round(pending_year_total, 2)
        period_values_dict['total'] = round(quarter_total, 2)
        period_values_dict['year_total'] = round(year_total, 2)

    if queryset3 and queryset4:
        period_values_sum = queryset3.aggregate(**period_values_aggregations)
        netperd_fields = [f"netperd{i}" for i in range(1, ytd_range[int(current_q)])]
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = queryset4.aggregate(
            **total_period_values_aggregations)
        quarter_total_actual = sum(period_values_sum.values())
        year_total_actual = sum(total_period_values_sum.values())

        period_values_dict['total_actual'] = round(quarter_total_actual, 2)
        period_values_dict['year_total_actual'] = round(year_total_actual, 2)
    total_actual = abs(period_values_dict['total_actual'])
    year_total_actual = abs(period_values_dict['year_total_actual'])
    pending_total = period_values_dict['pending']
    year_pending_total = period_values_dict['pending_year_total']

    available = round(Decimal(period_values_dict['total']), 2) - round(total_actual, 2) - round(
        pending_total, 2)
    year_available = round(Decimal(period_values_dict['available_year_total']), 2) - round(year_total_actual,
                                                                                           2) - round(
        year_pending_total, 2)
    period_values_dict['available'] = "{:,.2f}".format(available)
    period_values_dict['available_year_total'] = "{:,.2f}".format(year_available)

    return period_values_dict


def netperd_aggregate(queryset, queryset2, queryset3, queryset4, netperd, dept_id):
    """
    Calculate budget-related values for a specific period.

    Parameters:
    - queryset: Budget totals queryset for specific period
    - queryset2: Budget totals queryset for year to date
    - queryset3: Glafs  queryset for specific period
    - queryset4: Glafs queryset for year to date
    - netperd: Period identifier

    Returns:
    - Dictionary containing calculated values
    """
    year = FinancialYear.objects.get(is_active=True)

    period_values_dict = {
        'total': Decimal('0.00'),
        'total_actual': Decimal('0.00'),
        'available': Decimal('0.00'),
        'pending': Decimal('0.00'),
        'year_total': Decimal('0.00'),
        'year_total_actual': Decimal('0.00'),
        'pending_year_total': Decimal('0.00'),
        'available_year_total': Decimal('0.00')
    }

    ranges = {
        'netperd1': f'audtdate__gte={year.year}0101,audtdate__lte={year.year}0131',
        'netperd2': f'audtdate__gte={year.year}0201,audtdate__lte={year.year}0229',
        'netperd3': f'audtdate__gte={year.year}0301,audtdate__lte={year.year}0331',
        'netperd4': f'audtdate__gte={year.year}0401,audtdate__lte={year.year}0430',
        'netperd5': f'audtdate__gte={year.year}0501,audtdate__lte={year.year}0531',
        'netperd6': f'audtdate__gte={year.year}0601,audtdate__lte={year.year}0630',
        'netperd7': f'audtdate__gte={year.year}0701,audtdate__lte={year.year}0731',
        'netperd8': f'audtdate__gte={year.year}0801,audtdate__lte={year.year}0831',
        'netperd9': f'audtdate__gte={year.year}0901,audtdate__lte={year.year}0930',
        'netperd10': f'audtdate__gte={year.year}1001,audtdate__lte={year.year}1031',
        'netperd11': f'audtdate__gte={year.year}1101,audtdate__lte={year.year}1130',
        'netperd12': f'audtdate__gte={year.year}1201,audtdate__lte={year.year}1231',
    }

    netperd_fields = [f"netperd{i}" for i in range(1, int(netperd.replace('netperd', '')) + 1)]

    if queryset and queryset2:
        period_aggregation = queryset.aggregate(Sum(netperd)).get(f'{netperd}__sum', 0)
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = queryset2.aggregate(**total_period_values_aggregations)
        year_total_actual = sum(total_period_values_sum.values())

        start_date, end_date = ranges[netperd].split(',')
        start_date = int(start_date.split('=')[1])
        end_date = int(end_date.split('=')[1])
        acctids = [obj['account__acctfmttd'] for obj in queryset]
        q1 = reduce(or_, [Q(idglacct__icontains=item) for item in acctids])
        q2 = reduce(or_, [Q(glacct__icontains=item) for item in acctids])
        q3 = reduce(or_, [Q(glnonstkcr__icontains=item) for item in acctids])

        pending_expenses = Enebd.objects.using('sql_server').filter(
            q1,
            status=1,
            expdate__gte=start_date,
            expdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0
        print(pending_expenses)
        pending_expenses_all = Enebd.objects.using('sql_server').filter(
            q1,
            status=1,
            expdate__gte=int(year.year + '0101'),
            expdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_requisition = Enrqnl.objects.using('sql_server').filter(
            q2,
            status=1,
            exparrival__gte=start_date,
            exparrival__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0
        print(pending_requisition)
        pending_requisition_all = Enrqnl.objects.using('sql_server').filter(
            q2,
            status=1,
            audtdate__gte=int(year.year + '0101'),
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order = Poporl.objects.using('sql_server').filter(
            q3,
            completion=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0
        print(pending_purchase_order)
        pending_purchase_order_all = Poporl.objects.using('sql_server').filter(
            q3,
            completion=1,
            audtdate__gte=int(year.year + '0101'),
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_total = pending_expenses + pending_requisition + pending_purchase_order
        pending_year_total = pending_expenses_all + pending_requisition_all + pending_purchase_order_all

        period_values_dict['pending'] = round(pending_total, 2)
        period_values_dict['pending_year_total'] = round(pending_year_total, 2)
        period_values_dict['available'] = round(period_aggregation, 2) - round(pending_total, 2)
        period_values_dict['available_year_total'] = round(year_total_actual, 2) - round(pending_year_total, 2)
        period_values_dict['total'] = round(period_aggregation, 2)
        period_values_dict['year_total'] = round(year_total_actual, 2)

    if queryset3 and queryset4:

        period_aggregation_actual = 0
        for obj in queryset3:
            period_aggregation_actual += obj[f'{netperd}']

        # Calculate the total period values sum in queryset4
        total_period_values_sum_actual = 0
        for period in netperd_fields:
            total_period_values_sum_actual += queryset4.aggregate(Sum(period))[f'{period}__sum'] or 0

        year_total_actual = total_period_values_sum_actual

        period_values_dict['total_actual'] = round(period_aggregation_actual, 2)
        period_values_dict['year_total_actual'] = round(year_total_actual, 2)
    total_actual = abs(Decimal(period_values_dict['total_actual']))
    year_total_actual = abs(Decimal(period_values_dict['year_total_actual']))
    pending_total = Decimal(period_values_dict['pending'])
    year_pending_total = Decimal(period_values_dict['pending_year_total'])

    available = round(Decimal(period_values_dict['total']), 2) - round(total_actual, 2) - round(
        pending_total, 2)
    year_available = round(Decimal(period_values_dict['year_total']), 2) - round(year_total_actual,
                                                                                    2) - round(
        year_pending_total, 2)
    period_values_dict['available'] = "{:,.2f}".format(available)
    period_values_dict['available_year_total'] = "{:,.2f}".format(year_available)
    return period_values_dict

def process_department(args):
                dept_name, dept_id, queryset, queryset2, queryset3, queryset4, current_q = args
                values = netperd_aggregate(queryset, queryset3, queryset2, queryset4, current_q, dept_id)
                return {
                    'acctdesc': dept_name,
                    **values,
                    'id': dept_id
                }
def q1_values(request):
    """
    used to display budget,actual and pending information to budget managers i.e users with role 002
    :param request:
    :return: html template
    """
    if request.user.is_authenticated:
        now = datetime.now()
        year = FinancialYear.objects.get(is_active=True)
        budget_set = 'Budget 1'  # necessary because there can be upto 5 budget variations and end user has choice among them
        if 'budget_set' in request.GET:
            budget_set = request.GET['budget_set']
        exclude_zero_perds = Q()

        capex_netper_sum = Decimal('0')
        opex_netper_sum = Decimal('0')
        department_conditions = Q()
        for department_id in range(1, 15):  # Assuming departments IDs range from 1 to 12
            department_conditions |= Q(department_id=department_id)
        budget_totals_all = BudgetTotals.objects.filter(
            Q(posted=False) & department_conditions & Q(budget_set=budget_set) & Q(year=year.year)
            # Base budget totals queryset
        ).order_by('-department_id', '-account_id')
        capex_budget_totals_all = budget_totals_all.filter(department_id=13)
        opex_budget_totals_all = budget_totals_all.exclude(department_id=13)
        opex_total_sum = budget_totals_all.exclude(department_id__in=[13, 15, 16]).aggregate(total_sum=Sum('total'))[
                             'total_sum'] or 0

        capex_total_sum = budget_totals_all.filter(department_id=13).aggregate(total_sum=Sum('total'))['total_sum'] or 0

        opex_actual_total = Glafs.objects.using('sql_server').filter(
            acctid__in=[obj.account.acctid for obj in opex_budget_totals_all], fscsdsg='A', fscscurn='ZMW',
            curntype='F', fscsyr=year.year)
        capex_actual_total = Glafs.objects.using('sql_server').filter(
            acctid__in=[obj.account.acctid for obj in capex_budget_totals_all], fscsdsg='A', fscscurn='ZMW',
            curntype='F', fscsyr=year.year)

        for total in opex_actual_total:
            opex_netper_sum += total.get_netper_sum()

        for total in capex_actual_total:
            capex_netper_sum += total.get_netper_sum()
        if 'quarter' in request.GET:
            current_q = request.GET['quarter']
            months_in_current_quarter = netperd_in_quarter(int(current_q))
            net_perds = [month for month in months_in_current_quarter]
            for net_perd in net_perds:
                exclude_zero_perds |= Q(**{f"{net_perd}": 0})
            budget_totals = BudgetTotals.objects.filter(
                Q(posted=False) & department_conditions & Q(budget_set=budget_set) & Q(year=year.year)
                ).order_by('-department_id', '-account_id').values(
                *net_perds, 'account_id', 'account__acctdesc', 'account__acctfmttd')

            acctIdsCeo = [obj['account_id'] for obj in budget_totals.filter(department_id=1)]
            acctIdsIa = [obj['account_id'] for obj in budget_totals.filter(department_id=2)]
            acctIdsSc = [obj['account_id'] for obj in budget_totals.filter(department_id=3)]
            acctIdsBds = [obj['account_id'] for obj in budget_totals.filter(department_id=4)]
            acctIdsPr = [obj['account_id'] for obj in budget_totals.filter(department_id=5)]
            acctIdsTech = [obj['account_id'] for obj in budget_totals.filter(department_id=6)]
            acctIdsIs = [obj['account_id'] for obj in budget_totals.filter(department_id=7)]
            acctIdsLr = [obj['account_id'] for obj in budget_totals.filter(department_id=8)]
            acctIdsHc = [obj['account_id'] for obj in budget_totals.filter(department_id=9)]
            acctIdsSm = [obj['account_id'] for obj in budget_totals.filter(department_id=10)]
            acctIdsAdmin = [obj['account_id'] for obj in budget_totals.filter(department_id=11)]
            acctIdsFin = [obj['account_id'] for obj in budget_totals.filter(department_id=12)]
            acctIdsAsset = [obj['account_id'] for obj in budget_totals.filter(department_id=13)]
            acctIdsIncomeTowers = [obj['account_id'] for obj in
                                   budget_totals.filter(department_id=14)]

            ceo_queryset = budget_totals.filter(department_id=1)
            ceo_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsCeo, fscsdsg='A', fscscurn='ZMW',
                                                                   curntype='F',
                                                                   fscsyr=year.year).values(
                *net_perds,
                'acctid')
            internal_audit_queryset = budget_totals.filter(department_id=2)
            internal_audit_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsIa, fscsdsg='A',
                                                                              fscscurn='ZMW',
                                                                              curntype='F', fscsyr=year.year).values(
                *net_perds, 'acctid')
            supply_chain_queryset = budget_totals.filter(department_id=3)
            supply_chain_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsSc, fscsdsg='A',
                                                                            fscscurn='ZMW',
                                                                            curntype='F', fscsyr=year.year).values(
                *net_perds, 'acctid')
            bds_queryset = budget_totals.filter(department_id=4)
            bds_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsBds, fscsdsg='A', fscscurn='ZMW',
                                                                   curntype='F',
                                                                   fscsyr=year.year).values(
                *net_perds,
                'acctid')
            public_relations_queryset = budget_totals.filter(department_id=5)
            public_relations_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsPr, fscsdsg='A',
                                                                                fscscurn='ZMW',
                                                                                curntype='F', fscsyr=year.year).values(
                *net_perds, 'acctid')
            technical_queryset = budget_totals.filter(department_id=6)
            technical_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsTech, fscsdsg='A',
                                                                         fscscurn='ZMW',
                                                                         curntype='F', fscsyr=year.year).values(
                *net_perds, 'acctid')
            information_systems_queryset = budget_totals.filter(department_id=7)
            information_systems_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsIs, fscsdsg='A',
                                                                                   fscscurn='ZMW',
                                                                                   curntype='F',
                                                                                   fscsyr=year.year).values(
                *net_perds,
                'acctid')
            legal_risk_queryset = budget_totals.filter(department_id=8)
            legal_risk_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsLr, fscsdsg='A',
                                                                          fscscurn='ZMW',
                                                                          curntype='F', fscsyr=year.year).values(
                *net_perds, 'acctid')
            human_capital_queryset = budget_totals.filter(department_id=9)
            human_capital_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsHc, fscsdsg='A',
                                                                             fscscurn='ZMW',
                                                                             curntype='F', fscsyr=year.year).values(
                *net_perds, 'acctid')
            sales_marketing_queryset = budget_totals.filter(department_id=10)
            sales_marketing_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsSm, fscsdsg='A',
                                                                               fscscurn='ZMW',
                                                                               curntype='F', fscsyr=year.year).values(
                *net_perds, 'acctid')
            admin_queryset = budget_totals.filter(department_id=11)
            admin_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsAdmin, fscsdsg='A',
                                                                     fscscurn='ZMW', curntype='F',
                                                                     fscsyr=year.year).values(
                *net_perds, 'acctid')
            finance_queryset = budget_totals.filter(department_id=12)
            finance_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsFin, fscsdsg='A',
                                                                       fscscurn='ZMW', curntype='F',
                                                                       fscsyr=year.year).values(
                *net_perds, 'acctid')
            asset_queryset = budget_totals.filter(department_id=13)
            asset_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsAsset, fscsdsg='A',
                                                                     fscscurn='ZMW', curntype='F',
                                                                     fscsyr=year.year).values(
                *net_perds, 'acctid')
            incometw_queryset = budget_totals.filter(department_id=14)
            incometw_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsIncomeTowers, fscsdsg='A',
                                                                        fscscurn='ZMW',
                                                                        curntype='F', fscsyr=year.year).values(
                *net_perds, 'acctid')

            querysets = [
                ('CEO', 1, ceo_queryset, ceo_actuals),
                ('Internal Audit', 2, internal_audit_queryset, internal_audit_actuals),
                ('Supply Chain', 3, supply_chain_queryset, supply_chain_actuals),
                ('BDS', 4, bds_queryset, bds_actuals),
                ('Public Relations', 5, public_relations_queryset, public_relations_actuals),
                ('Technical', 6, technical_queryset, technical_actuals),
                ('Information Systems', 7, information_systems_queryset, information_systems_actuals),
                ('Legal & Risk', 8, legal_risk_queryset, legal_risk_actuals),
                ('Human Capital', 9, human_capital_queryset, human_capital_actuals),
                ('Sales & Marketing', 10, sales_marketing_queryset, sales_marketing_actuals),
                ('Administration', 11, admin_queryset, admin_actuals),
                ('Finance', 12, finance_queryset, finance_actuals),
                ('Staff Remuneration & Related costs', 14, incometw_queryset, incometw_actuals),

                ('CAPEX', 13, asset_queryset, asset_actuals),

            ]
            unit_totals = []
            before = datetime.now()
            for dept_name, dept_id, queryset, queryset2 in querysets:
                actuals = Glafs.objects.using('sql_server').filter(
                    acctid__in=[obj['account_id'] for obj in queryset], fscsdsg='A', fscscurn='ZMW', curntype='F',
                    fscsyr=year.year)
                values = quarter_aggregate(queryset, budget_totals_all.filter(department_id=dept_id), queryset2,
                                           actuals, current_q, dept_id)
                unit_totals.append(
                    {
                        'acctdesc': dept_name,
                        **values,
                        'id': dept_id
                    }
                )
            data = {'dept_totals': unit_totals}
            opex_total = opex_total_actual = opex_available = opex_pending = opex_ytd_total = opex_ytd_total_actual = opex_ytd_available = opex_ytd_pending = Decimal(
                '0.00')
            capex_total = capex_total_actual = capex_available = capex_pending = capex_ytd_total = capex_ytd_total_actual = capex_ytd_available = capex_ytd_pending = Decimal(
                '0.00')
            for item in data['dept_totals']:
                if item['id'] != 13:
                    opex_total += item['total']
                    opex_total_actual += item['total_actual']
                    opex_available += Decimal(item['available'].replace(',', ''))
                    opex_pending += item['pending']
                    opex_ytd_total += item['year_total']
                    opex_ytd_total_actual += item['year_total_actual']
                    opex_ytd_available += Decimal(item['available_year_total'].replace(',', ''))
                    opex_ytd_pending += item['pending_year_total']
                elif item['id'] == 13:
                    capex_total = item['total']
                    capex_total_actual = item['total_actual']
                    capex_available = Decimal(item['available'].replace(',', ''))
                    capex_pending = item['pending']
                    capex_ytd_total = item['year_total']
                    capex_ytd_total_actual = item['year_total_actual']
                    capex_ytd_available = Decimal(item['available_year_total'].replace(',', ''))
                    capex_ytd_pending = item['pending_year_total']

            context = {'data': data, 'period': f'Q{current_q}',
                       'capex_total': round(capex_total_sum, 2),
                       'capex_actual_total': round(capex_netper_sum, 2),
                       'opex_total': opex_total_sum,
                       'opex_actual_total': round(opex_netper_sum, 2),
                       'opex_period_total': opex_total,
                       'opex_period_actual_total': opex_total_actual,
                       'opex_period_pending_total': opex_pending,
                       'opex_period_available_total': opex_available,
                       'opex_ytd_total': opex_ytd_total,
                       'opex_ytd_actual_total': opex_ytd_total_actual,
                       'opex_ytd_pending_total': opex_ytd_pending,
                       'opex_ytd_available_total': opex_ytd_available,
                       'period_total': capex_total,
                       'period_actual_total': capex_total_actual,
                       'period_pending_total': capex_pending,
                       'period_available_total': capex_available,
                       'ytd_total': capex_ytd_total,
                       'ytd_actual_total': capex_ytd_total_actual,
                       'ytd_pending_total': capex_ytd_pending,
                       'ytd_available_total': capex_ytd_available,
                        'year':year
                       }

            return render(request, 'index.html', context)



        else:
            current_q = current_period()
            if 'period' in request.GET:
                current_q = request.GET['period']

            budget_totals = BudgetTotals.objects.filter(
                Q(posted=False) & department_conditions & Q(budget_set=budget_set) & Q(year=year.year)
                ).order_by('-department_id', '-account_id').values(
                current_q, 'account_id', 'account__acctdesc', 'account__acctfmttd')
            acctIdsCeo = [obj['account_id'] for obj in budget_totals.filter(department_id=1)]
            acctIdsIa = [obj['account_id'] for obj in budget_totals.filter(department_id=2)]
            acctIdsSc = [obj['account_id'] for obj in budget_totals.filter(department_id=3)]
            acctIdsBds = [obj['account_id'] for obj in budget_totals.filter(department_id=4)]
            acctIdsPr = [obj['account_id'] for obj in budget_totals.filter(department_id=5)]
            acctIdsTech = [obj['account_id'] for obj in budget_totals.filter(department_id=6)]
            acctIdsIs = [obj['account_id'] for obj in budget_totals.filter(department_id=7)]
            acctIdsLr = [obj['account_id'] for obj in budget_totals.filter(department_id=8)]
            acctIdsHc = [obj['account_id'] for obj in budget_totals.filter(department_id=9)]
            acctIdsSm = [obj['account_id'] for obj in budget_totals.filter(department_id=10)]
            acctIdsAdmin = [obj['account_id'] for obj in budget_totals.filter(department_id=11)]
            acctIdsFin = [obj['account_id'] for obj in budget_totals.filter(department_id=12)]
            acctIdsAsset = [obj['account_id'] for obj in budget_totals.filter(department_id=13)]
            acctIdsIncomeTowers = [obj['account_id'] for obj in budget_totals.filter(department_id=14)]
            exclude_filter = {f'{current_q}': 0}
            ceo_queryset = budget_totals.filter(department_id=1)
            ceo_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsCeo, fscsdsg='A', fscscurn='ZMW',
                                                                   curntype='F', fscsyr=year.year).values(current_q,

                                                                                                            'acctid')
            ceo_ytd = budget_totals_all.filter(department_id=1)
            ceo_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsCeo, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #Internal Audit
            internal_audit_queryset = budget_totals.filter(department_id=2)
            internal_audit_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsIa, fscsdsg='A',
                                                                              fscscurn='ZMW',
                                                                              curntype='F', fscsyr=year.year).values(
                current_q, 'acctid')
            internal_audit_ytd =budget_totals_all.filter(department_id=2)
            internal_audit_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsIa, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #supply chain
            supply_chain_queryset = budget_totals.filter(department_id=3)
            supply_chain_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsSc, fscsdsg='A',
                                                                            fscscurn='ZMW',
                                                                            curntype='F', fscsyr=year.year).values(
                current_q, 'acctid')
            supply_chain_ytd =budget_totals_all.filter(department_id=3)
            supply_chain_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsSc, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #bds
            bds_queryset = budget_totals.filter(department_id=4)
            bds_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsBds, fscsdsg='A', fscscurn='ZMW',
                                                                   curntype='F',
                                                                   fscsyr=year.year).values(
                current_q,
                'acctid')
            bds_ytd =budget_totals_all.filter(department_id=4)
            bds_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsBds, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #public relations
            public_relations_queryset = budget_totals.filter(department_id=5)
            public_relations_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsPr, fscsdsg='A',
                                                                                fscscurn='ZMW', curntype='F',
                                                                                fscsyr=year.year).values(
                current_q, 'acctid')
            public_relations_ytd =budget_totals_all.filter(department_id=5)
            public_relations_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsPr, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #tech
            technical_queryset = budget_totals.filter(department_id=6)
            technical_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsTech, fscsdsg='A',
                                                                         fscscurn='ZMW',
                                                                         curntype='F', fscsyr=year.year).values(
                current_q, 'acctid')
            technical_ytd = budget_totals_all.filter(department_id=6)
            technical_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsTech, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #information systems
            information_systems_queryset = budget_totals.filter(department_id=7)
            information_systems_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsIs, fscsdsg='A',
                                                                                   fscscurn='ZMW', curntype='F',
                                                                                   fscsyr=year.year).values(current_q,
                                                                                                            'acctid')
            information_systems_ytd = budget_totals_all.filter(department_id=7)
            information_systems_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsIs, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #legal_risk
            legal_risk_queryset = budget_totals.filter(department_id=8)
            legal_risk_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsLr, fscsdsg='A',
                                                                          fscscurn='ZMW',
                                                                          curntype='F', fscsyr=year.year).values(
                current_q, 'acctid')
            legal_risk_ytd =budget_totals_all.filter(department_id=8)
            legal_risk_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsLr, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #Human_capital
            human_capital_queryset = budget_totals.filter(department_id=9)
            human_capital_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsHc, fscsdsg='A',
                                                                             fscscurn='ZMW',
                                                                             curntype='F', fscsyr=year.year).values(
                current_q, 'acctid')
            human_capital_ytd = budget_totals_all.filter(department_id=9)
            human_capital_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsHc, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #sales_marketing
            sales_marketing_queryset = budget_totals.filter(department_id=10)
            sales_marketing_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsSm, fscsdsg='A',
                                                                               fscscurn='ZMW', curntype='F',
                                                                               fscsyr=year.year).values(
                current_q, 'acctid')
            sales_marketing_ytd =budget_totals_all.filter(department_id=10)
            sales_marketing_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsSm, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #admin
            admin_queryset = budget_totals.filter(department_id=11)
            admin_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsAdmin, fscsdsg='A',
                                                                     fscscurn='ZMW', curntype='F',
                                                                     fscsyr=year.year).values(
                current_q, 'acctid')
            admin_ytd =budget_totals_all.filter(department_id=11)
            admin_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsAdmin, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #finance
            finance_queryset = budget_totals.filter(department_id=12)
            finance_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsFin, fscsdsg='A',
                                                                       fscscurn='ZMW', curntype='F',
                                                                       fscsyr=year.year).values(
                current_q, 'acctid')
            finance_ytd = budget_totals_all.filter(department_id=12)
            finance_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsFin, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #asset
            asset_queryset = budget_totals.filter(department_id=13)
            asset_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsAsset, fscsdsg='A',
                                                                     fscscurn='ZMW', curntype='F',
                                                                     fscsyr=year.year).values(
                current_q, 'acctid')
            asset_ytd =budget_totals_all.filter(department_id=13)
            asset_ytd_actuals = Glafs.objects.using('sql_server').filter(
                acctid__in=acctIdsAsset, fscsdsg='A', fscscurn='ZMW', curntype='F',
                fscsyr=year.year)
            #Staff and Remunerations
            incometw_queryset = budget_totals.filter(department_id=14)
            incometw_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsIncomeTowers, fscsdsg='A',
                                                                        fscscurn='ZMW', curntype='F',
                                                                        fscsyr=year.year).values(
                current_q, 'acctid')

            incometw_ytd =budget_totals_all.filter(department_id=14)
            income_ytd_actuals = Glafs.objects.using('sql_server').filter(
                    acctid__in=acctIdsIncomeTowers, fscsdsg='A', fscscurn='ZMW', curntype='F',
                    fscsyr=year.year)


            querysets = [
                ('CEO', 1, ceo_queryset, ceo_actuals,ceo_ytd,ceo_ytd_actuals),
               ('Internal Audit', 2, internal_audit_queryset, internal_audit_actuals,internal_audit_ytd,internal_audit_ytd_actuals),
                ('Supply Chain', 3, supply_chain_queryset, supply_chain_actuals,supply_chain_ytd,supply_chain_ytd_actuals),
                ('BDS', 4, bds_queryset, bds_actuals,bds_ytd,bds_ytd_actuals),
                ('Public Relations', 5, public_relations_queryset, public_relations_actuals,public_relations_ytd,public_relations_ytd_actuals),
                ('Technical', 6, technical_queryset, technical_actuals,technical_ytd,technical_ytd_actuals),
               ('Information Systems', 7, information_systems_queryset, information_systems_actuals, information_systems_ytd,information_systems_ytd_actuals),
                ('Legal & Risk', 8, legal_risk_queryset, legal_risk_actuals, legal_risk_ytd,legal_risk_ytd_actuals),
                ('Human Capital', 9, human_capital_queryset, human_capital_actuals, human_capital_ytd,human_capital_ytd_actuals),
                ('Sales & Marketing', 10, sales_marketing_queryset, sales_marketing_actuals,sales_marketing_ytd,sales_marketing_ytd_actuals),
                ('Administration', 11, admin_queryset, admin_actuals,admin_ytd,admin_ytd_actuals),
               ('Finance', 12, finance_queryset, finance_actuals, finance_ytd,finance_ytd_actuals),
                ('Staff Remuneration & Related costs', 14, incometw_queryset, incometw_actuals, incometw_ytd,income_ytd_actuals),

                ('CAPEX', 13, asset_queryset, asset_actuals,asset_ytd,asset_ytd_actuals),

            ]
            unit_totals = []

            for dept_name, dept_id, queryset, queryset2,queryset3,queryset4 in querysets:
                
                values = netperd_aggregate(queryset, queryset3, queryset2,queryset4, current_q, dept_id)
                unit_totals.append(
                    {
                        'acctdesc': dept_name,
                        **values,
                        'id': dept_id
                    }
                )


            data = {'dept_totals': unit_totals}
            print(data['dept_totals'])

            ranges = {
                'netperd1': 'January',
                'netperd2': 'February',
                'netperd3': 'March',
                'netperd4': 'April',
                'netperd5': 'May',
                'netperd6': 'June',
                'netperd7': 'July',
                'netperd8': 'August',
                'netperd9': 'September',
                'netperd10': 'October',
                'netperd11': 'November',
                'netperd12': 'December'
            }
            opex_total = opex_total_actual = opex_available = opex_pending = opex_ytd_total = opex_ytd_total_actual = opex_ytd_available = opex_ytd_pending = Decimal(
                '0.00')
            capex_total = capex_total_actual = capex_available = capex_pending = capex_ytd_total = capex_ytd_total_actual = capex_ytd_available = capex_ytd_pending = Decimal(
                '0.00')

            for item in data['dept_totals']:
                if item['id'] != 13:
                    opex_total += item['total']
                    opex_total_actual += item['total_actual']
                    opex_available += Decimal(item['available'].replace(',', ''))
                    opex_pending += item['pending']
                    opex_ytd_total += item['year_total']
                    opex_ytd_total_actual += item['year_total_actual']
                    opex_ytd_available += Decimal(item['available_year_total'].replace(',', ''))
                    opex_ytd_pending += item['pending_year_total']
                elif item['id'] == 13:
                    capex_total = item['total']
                    capex_total_actual = item['total_actual']
                    capex_available = Decimal(item['available'].replace(',', ''))
                    capex_pending = item['pending']
                    capex_ytd_total = item['year_total']
                    capex_ytd_total_actual = item['year_total_actual']
                    capex_ytd_available = Decimal(item['available_year_total'].replace(',', ''))
                    capex_ytd_pending = item['pending_year_total']

            context = {'data': data, 'period': ranges[current_q],
                       'capex_total': round(capex_total_sum, 2),
                       'capex_actual_total': round(capex_netper_sum, 2),
                       'opex_total': opex_total_sum,
                       'opex_actual_total': round(opex_netper_sum, 2),
                       'opex_period_total': opex_total,
                       'opex_period_actual_total': opex_total_actual,
                       'opex_period_pending_total': opex_pending,
                       'opex_period_available_total': opex_available,
                       'opex_ytd_total': opex_ytd_total,
                       'opex_ytd_actual_total': opex_ytd_total_actual,
                       'opex_ytd_pending_total': opex_ytd_pending,
                       'opex_ytd_available_total': opex_ytd_available,
                       'period_total': capex_total,
                       'period_actual_total': capex_total_actual,
                       'period_pending_total': capex_pending,
                       'period_available_total': capex_available,
                       'ytd_total': capex_ytd_total,
                       'ytd_actual_total': capex_ytd_total_actual,
                       'ytd_pending_total': capex_ytd_pending,
                       'ytd_available_total': capex_ytd_available,
                       'year': year
                       }

            return render(request, 'index.html', context)

def toggle_log(request):
    year = FinancialYear.objects.get(is_active=True)
    changelog_instance = ChangeLog.objects.get(year=year.year)  # Fetch the instance you want to toggle
    changelog_instance.toggle_flag()
    return redirect('budgets:settings')

def search_dashboard(request):
    year = FinancialYear.objects.get(is_active=True)
    period = request.GET['period']
    budget_set = 'Budget 1'
    if 'budget_set' in request.GET:
        budget_set = request.GET['budget_set']

    department = request.GET.get('department', None)
    if department:
        field_mapping = {
            'account_id': 'account__acctid__icontains',
            'account_name': 'account__acctdesc__icontains',
        }

        filter = request.GET.get('filter')
        value = request.GET.get('value')
        dept_totals = []
        if 'netperd' in period:

            budget_total = BudgetTotals.objects.filter(
                Q(**{field_mapping[filter]: value}, department_id=department) & Q(year=year.year) & Q(
                    budget_set=budget_set)).values(period, 'id', 'account_id', 'department__name', 'account__acctdesc',
                                                   'account__acctfmttd').order_by(
                '-department_id', '-account_id')
            budget_total_all = BudgetTotals.objects.filter(Q(**{field_mapping[filter]: value}, )).order_by(
                '-department_id', '-account_id')

            for totals, totals2 in zip(budget_total, budget_total_all):
                dept_totals.append({
                    'acctdesc': totals['account__acctdesc'],
                    'acctid': totals['account_id'],

                    'id': totals['department__name'],
                    'objId': totals['id'],
                    'table_type': 'quarter'
                })
            data = {'dept_totals': dept_totals}
            print(data)
            return JsonResponse({'data': data}, status=200)
    else:
        field_mapping = {
            'account_id': 'account__acctid__icontains',
            'account_name': 'account__acctdesc__icontains',
        }

        filter = request.GET.get('filter')
        value = request.GET.get('value')
        dept_totals = []
        if 'netperd' in period:

            budget_total = BudgetTotals.objects.filter(
                Q(**{field_mapping[filter]: value}) & Q(year=year.year) & Q(budget_set=budget_set)).values(period, 'id',
                                                                                                           'account_id',
                                                                                                           'department__name',
                                                                                                           'account__acctdesc',
                                                                                                           'account__acctfmttd').order_by(
                '-department_id', '-account_id')
            budget_total_all = BudgetTotals.objects.filter(
                Q(**{field_mapping[filter]: value}) & Q(year=year.year) & Q(budget_set=budget_set)).order_by(
                '-department_id', '-account_id')

            for totals, totals2 in zip(budget_total, budget_total_all):
                dept_totals.append({
                    'acctdesc': totals['account__acctdesc'],
                    'acctid': totals['account_id'],

                    'id': totals['department__name'],
                    'objId': totals['id'],
                    'table_type': 'quarter'
                })
            data = {'dept_totals': dept_totals}
            return JsonResponse({'data': data}, status=200)


def load_status(request):
    status = BudgetStatus.objects.all()
    for obj in status:
        BudgetStatus.objects.create(
            budget_set=obj.budget_set,
            year='2026',
            department=obj.department,
            is_active=False,
            is_complete=obj.is_complete,
            updated_by=obj.updated_by,
            comment=obj.comment
        )


def load_budget(request):
    budget_instance = BudgetTotals.objects.all()
    budget_accounts = [obj.account_id for obj in budget_instance]
    accounts_instance = Accounts.objects.all()
    for item in accounts_instance:
        try:
            BudgetTotals.objects.create(
                account_id=item.acctid,  # An instance of the Accounts model
                currency_id=1,  # Replace with an actual Currency instance
                department_id=item.department_id,  # Replace with an actual Department instance
                year='2025',
                total=0.0,
                netperd1=0.0,
                netperd2=0.0,
                netperd3=0.0,
                netperd4=0.0,
                netperd5=0.0,
                netperd6=0.0,
                netperd7=0.0,
                netperd8=0.0,
                netperd9=0.0,
                netperd10=0.0,
                netperd11=0.0,
                netperd12=0.0,
                budget_set='Budget 2',  # Replace with a valid choice from SET_TYPE
                last_updated_by=request.user,  # Replace with an actual Users instance
                posted=False,
                exchange_rate=1.00)
            print(f'BudgetTotals created for account {item.acctid}')
        except Accounts.DoesNotExist:
            print(f'Account {item.acctid} does not exist in Accounts model')
        except Exception as e:
            print(f'An error occurred: {e}')

def load_lines_excel(request):


    # Load the Excel file
    file_path = 'sorted_dataset.xlsx'
    df = pd.read_excel(file_path)

    # Iterate through the rows of the Excel sheet
    for index, row in df.iterrows():
        # Get the account_id from the Excel sheet and ensure it is 45 characters long by adding trailing spaces
        account_id = str(row['account_id']).ljust(45)

        try:
            # Find the Accounts object, match with padded account_id
            account = Accounts.objects.get(acctid=account_id)

            # Find the corresponding BudgetTotals object
            budget_total = BudgetTotals.objects.get(account=account,budget_set="Budget 1", year= "2024")

            # Update the netperd fields and total
            budget_total.netperd1 = row['netperd1']
            budget_total.netperd2 = row['netperd2']
            budget_total.netperd3 = row['netperd3']
            budget_total.netperd4 = row['netperd4']
            budget_total.netperd5 = row['netperd5']
            budget_total.netperd6 = row['netperd6']
            budget_total.netperd7 = row['netperd7']
            budget_total.netperd8 = row['netperd8']
            budget_total.netperd9 = row['netperd9']
            budget_total.netperd10 = row['netperd10']
            budget_total.netperd11 = row['netperd11']
            budget_total.netperd12 = row['netperd12']
            budget_total.total = row['total']

            # Save the updated record
            budget_total.save()

        except BudgetTotals.DoesNotExist:
            print(f"BudgetTotals record not found for account_id: {account_id}")
        except Accounts.DoesNotExist:
            print(f"Account not found for account_id: {account_id}")

def load_accounts(request):

    accounts_instance = Accounts.objects.all()
    accts = [obj.acctid for obj in accounts_instance]
    glamf = Glamf.objects.using('sql_server').filter(acctid__icontains='71494120')
    for g in glamf:
        Accounts.objects.create(
            acctid=g.acctid,
            audtdate=g.audtdate,
            audttime=g.audttime,
            audtuser=g.audtuser,
            audtorg=g.audtorg,
            createdate=g.createdate,
            acctdesc=g.acctdesc,
            accttype=g.accttype,
            acctbal=g.acctbal,
            activesw=g.activesw,
            consldsw=g.consldsw,
            qtysw=g.qtysw,
            uom=g.uom,
            allocsw=g.allocsw,
            acctofset=g.acctofset,
            acctsrty=g.acctsrty,
            mcsw=g.mcsw,
            specsw=g.specsw,
            acctgrpcod=g.acctgrpcod,
            ctrlacctsw=g.ctrlacctsw,
            srceldgid=g.srceldgid,
            alloctot=g.alloctot,
            abrkid=g.abrkid,
            yracctclos=g.yracctclos,
            acctfmttd=g.acctfmttd,
            acsegval01=g.acsegval01,
            acsegval02=g.acsegval02,
            acsegval03=g.acsegval03,
            acsegval04=g.acsegval04,
            acsegval05=g.acsegval05,
            acsegval06=g.acsegval06,
            acsegval07=g.acsegval07,
            acsegval08=g.acsegval08,
            acsegval09=g.acsegval09,
            acsegval10=g.acsegval10,
            acctsegval=g.acctsegval,
            acctgrpscd=g.acctgrpscd,
            postosegid=g.postosegid,
            defcurncod=g.defcurncod,
            ovalues=g.ovalues,
            tovalues=g.tovalues,
            rollupsw=g.rollupsw,
            department_id=1
        )
def load_lines(request):
    year = FinancialYear.objects.filter(is_active=True)
    glafs_instance_main = Glafs.objects.using('sql_server').filter(fscsdsg='1', fscscurn='ZMW', curntype='F',
                                                                  fscsyr=2025).all()
    gl_accts = [obj.acctid for obj in glafs_instance_main]
    accounts_instance = Accounts.objects.all()
    accts = [obj.acctid for obj in accounts_instance]

    budget_instance = BudgetTotals.objects.filter(budget_set="Budget 1", year="2025")
    budget_accts = [obj.account_id for obj in budget_instance]
    
    """for accounts in budget_instance:
            accounts.total = 0
            for i in range(1, 13):
                setattr(accounts, f'netperd{i}', 0)"""
    for obj1, obj2 in zip(accounts_instance, glafs_instance_main):

        print(obj2.acctid)
        acc = Accounts.objects.get(acctid=obj2.acctid)
        try:
            netperd_total = (
                    obj2.netperd1 + obj2.netperd2 + obj2.netperd3 + obj2.netperd4 +
                    obj2.netperd5 + obj2.netperd6 + obj2.netperd7 + obj2.netperd8 +
                    obj2.netperd9 + obj2.netperd10 + obj2.netperd11 + obj2.netperd12
            )
            BudgetLines.objects.create(
                currency_id =1,
                usage=0.0,
                rate=0.0,
                staff=0.0,
                factor=0.0,
                budget_set= "Budget 1",
                last_updated = datetime.now(),
                item_description="Opening Budget Line",
                department_id=acc.department_id,
                last_updated_by_id=2,
                exchange_rate=1.00,
                year="2025",
                total=netperd_total,
                netperd1=obj2.netperd1,
                netperd2=obj2.netperd2,
                netperd3=obj2.netperd3,
                netperd4=obj2.netperd4,
                netperd5=obj2.netperd5,
                netperd6=obj2.netperd6,
                netperd7=obj2.netperd7,
                netperd8=obj2.netperd8,
                netperd9=obj2.netperd9,
                netperd10=obj2.netperd10,
                netperd11=obj2.netperd11,
                netperd12=obj2.netperd12,
                account_id=acc.acctid,
                assumption_id=1
            )
            """BudgetTotals.objects.filter(account_id=obj2.acctid, budget_set="Budget 1",year="2025").update(

                total=netperd_total,
                netperd1=obj2.netperd1,
                netperd2=obj2.netperd2,
                netperd3=obj2.netperd3,
                netperd4=obj2.netperd4,
                netperd5=obj2.netperd5,
                netperd6=obj2.netperd6,
                netperd7=obj2.netperd7,
                netperd8=obj2.netperd8,
                netperd9=obj2.netperd9,
                netperd10=obj2.netperd10,
                netperd11=obj2.netperd11,
                netperd12=obj2.netperd12,

            )"""
            print(f'BudgetTotals created for account {obj2.acctid}')
        except Accounts.DoesNotExist:
            print(f'Account {obj2.acctid} does not exist in Accounts model')
        except Exception as e:
            print(f'An error occurred: {e}')


def search_results_dashboard(request, acct):
    year = FinancialYear.objects.get(is_active=True)
    month_field_mapping = {
        1: 'netperd1',
        2: 'netperd2',
        3: 'netperd3',
        4: 'netperd4',
        5: 'netperd5',
        6: 'netperd6',
        7: 'netperd7',
        8: 'netperd8',
        9: 'netperd9',
        10: 'netperd10',
        11: 'netperd11',
        12: 'netperd12',
    }
    month_names = {
        1: "January",
        2: "February",
        3: "March",
        4: "April",
        5: "May",
        6: "June",
        7: "July",
        8: "August",
        9: "September",
        10: "October",
        11: "November",
        12: "December"
    }
    dept_totals = []
    active = BudgetVariations.objects.get(is_active=True,year=year.year)
    desc = get_object_or_404(BudgetTotals, account__acctid__icontains=acct, budget_set=active.budget_set,
                             year=year.year)
    budget_total_all = BudgetTotals.objects.filter(account__acctid__icontains=acct, posted=False,
                                                   budget_set=active.budget_set, year=year.year).order_by(
        '-department_id', '-account_id')

    for month in range(1, 13):
        month_field = month_field_mapping.get(month, None)
        if month_field:
            budget_total = BudgetTotals.objects.filter(account__acctid__icontains=acct, posted=False,
                                                       budget_set=active.budget_set, year=year.year).values(month_field,
                                                                                                            'account_id',
                                                                                                            'account__acctdesc',
                                                                                                            'account__acctfmttd').order_by(
                '-department_id', '-account_id')
            month_name = month_names[month]
            for totals, totals2 in zip(budget_total, budget_total_all):
                dept_totals.append({
                    'month': month_name,
                    'month_id': month,
                    'acctdesc': totals2.account.acctdesc,
                    'acctid': totals2.account.acctid,
                    **period_aggregate_singular(totals, totals2, month_field),
                    'objId': totals2.id,
                })

    data = {'dept_totals': dept_totals}
    context = {
        'data': data, 'account': desc.account.acctdesc, 'accountId': desc.account.acctid, 'dept': desc.department.id

    }
    return render(request, 'dashboard-search-results.html', context)


def actual_dashboard(request, department):
    year = FinancialYear.objects.get(is_active=True)
    dept_totals = []
    current_q = request.GET['period']
    account = request.GET.get('account', None)
    active = BudgetVariations.objects.get(is_active=True,year=year.year)
    if account:
        budget_total = BudgetTotals.objects.filter(department_id=department, budget_set=active.budget_set,
                                                   year=year.year, account__acctid__icontains=account,
                                                   posted=False).values('account_id', 'account__acctdesc',
                                                                        'account__acctfmttd').order_by(
            '-department_id', '-account_id')

        acctIds = [obj['account_id'] for obj in budget_total.filter(department_id=department)]

        name = Department.objects.get(id=department)

        for acc in acctIds:

            actual_amounts = get_actual_amounts(acc, current_q)
            if actual_amounts:  # Check if any value in the dictionary is not empty
                dept_totals.append({
                    'data': actual_amounts
                })

        data = {'dept_totals': dept_totals}
        context = {
            'data': data, 'account': name.name

        }
        return render(request, 'actual.html', context)
    else:
        budget_total = BudgetTotals.objects.filter(department_id=department, budget_set=active.budget_set,
                                                   year=year.year, posted=False).values('account_id',
                                                                                        'account__acctdesc',
                                                                                        'account__acctfmttd').order_by(
            '-department_id', '-account_id')

        acctIds = [obj['account_id'] for obj in budget_total.filter(department_id=department)]

        name = Department.objects.get(id=department)

        for acc in acctIds:

            actual_amounts = get_actual_amounts(acc, current_q)
            if actual_amounts:  # Check if any value in the dictionary is not empty
                dept_totals.append({
                    'data': actual_amounts
                })

        data = {'dept_totals': dept_totals}
        context = {
            'data': data, 'account': name.name

        }
        return render(request, 'actual.html', context)


def pending_dashboard(request, department):
    year = FinancialYear.objects.get(is_active=True)
    dept_totals = []
    current_q = request.GET['period']
    active = BudgetVariations.objects.get(is_active=True,year=year.year)
    account = request.GET.get('account', None)
    if account:
        budget_total = BudgetTotals.objects.filter(department_id=department, budget_set=active.budget_set,
                                                   account__acctid__icontains=account,
                                                   posted=False, year=year.year).values('account_id',
                                                                                        'account__acctdesc',
                                                                                        'account__acctfmttd').order_by(
            '-department_id', '-account_id')

        acctIds = [obj['account__acctfmttd'] for obj in budget_total.filter(department_id=department)]

        name = Department.objects.get(id=department)

        for acc in acctIds:

            pending_amounts = get_pending_amounts(acc, current_q)
            if pending_amounts:  # Check if any value in the dictionary is not empty
                dept_totals.append({
                    'data': pending_amounts
                })

        data = {'dept_totals': dept_totals}
        context = {
            'data': data, 'account': name.name

        }
        return render(request, 'pending.html', context)
    else:
        budget_total = BudgetTotals.objects.filter(department_id=department, budget_set=active.budget_set, posted=False,
                                                   year=year.year).values('account_id', 'account__acctdesc',
                                                                          'account__acctfmttd').order_by(
            '-department_id', '-account_id')

        acctIds = [obj['account__acctfmttd'] for obj in budget_total.filter(department_id=department)]

        name = Department.objects.get(id=department)

        for acc in acctIds:

            pending_amounts = get_pending_amounts(acc, current_q)
            if pending_amounts:  # Check if any value in the dictionary is not empty
                dept_totals.append({
                    'data': pending_amounts
                })

        data = {'dept_totals': dept_totals}
        context = {
            'data': data, 'account': name.name

        }
        return render(request, 'pending.html', context)


def quarter_dashboard_index_department_table(request):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)
        dept_totals = []
        budget_set = 'Budget 1'
        if 'budget_set' in request.GET:
            budget_set = request.GET['budget_set']
        department = request.GET.get('department')
        ptype = request.GET.get('period_type')
        current_q = request.GET.get('current_q')
        if ptype == 'month':
            # Im using the populated account ids here for coherence(we dont want to hve accounts showing that havent yet been entered)

            budget_total = BudgetTotals.objects.filter(department_id=department, posted=False, budget_set=budget_set,
                                                       year=year.year).values(current_q, 'account_id',
                                                                              'account__acctdesc',
                                                                              'account__acctfmttd').order_by(
                '-department_id', '-account_id')
            budget_total_all = BudgetTotals.objects.filter(department_id=department, posted=False,
                                                           budget_set=budget_set, year=year.year).order_by(
                '-department_id', '-account_id')

            for totals, totals2 in zip(budget_total, budget_total_all):
                dept_totals.append({
                    'acctdesc': totals2.account.acctdesc,
                    'acctid': totals2.account.acctid,
                    **period_aggregate_singular(totals, totals2, current_q),
                    'id': department,
                    'objId': totals2.id,
                    'table_type': 'quarter'
                })
            data = {'dept_totals': dept_totals}
            return JsonResponse({'data': data}, status=200)
        elif ptype == 'quarter':
            # Im using the populated account ids here for coherence(we dont want to hve accounts showing that havent yet been entered)
            months_in_current_quarter = netperd_in_quarter(int(current_q))
            net_perds = [month for month in months_in_current_quarter]
            budget_total = BudgetTotals.objects.filter(department_id=department, posted=False, budget_set=budget_set,
                                                       year=year.year).values(*net_perds, 'account_id',
                                                                              'account__acctdesc',
                                                                              'account_acctfmttd').order_by(
                '-department_id', '-account_id')
            budget_total_all = BudgetTotals.objects.filter(department_id=department, posted=False,
                                                           budget_set=budget_set, year=year.year).order_by(
                '-department_id', '-account_id')

            for totals, totals2 in zip(budget_total, budget_total_all):
                dept_totals.append({
                    'acctdesc': totals2.account.acctdesc,
                    'acctid': totals2.account.acctid,
                    **quarter_aggregate_singular(totals, totals2, current_q),
                    'id': department,
                    'objId': totals2.id,
                    'table_type': 'quarter'
                })
            data = {'dept_totals': dept_totals}
            return JsonResponse({'data': data}, status=200)

    else:
        return redirect('budgets:login')

        # Calculate the totals for each netperd


def dashboard_index_department(request):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)
        department = request.user.department.id

        if request.user.is_authenticated:
            budget_set = 'Budget 1'
            if 'budget_set' in request.GET:
                budget_set = request.GET['budget_set']

            dept_netper_sum = Decimal('0')

            if department == 9:
                budget_totals_all = BudgetTotals.objects.filter(
                    Q(posted=False) & Q(department_id__in=(9, 14,)) & Q(year=year.year) & Q(
                        budget_set=budget_set)).order_by(
                    '-department_id',
                    '-account_id')

                department_total_sum = budget_totals_all.filter(department_id=9).aggregate(total_sum=Sum('total'))[
                                           'total_sum'] or 0

                department_actual_total = Glafs.objects.using('sql_server').filter(
                    acctid__in=[obj.account.acctid for obj in budget_totals_all.filter(department_id=9)])
                for total in department_actual_total:
                    dept_netper_sum += total.get_netper_sum()

                if 'quarter' in request.GET:
                    current_q = request.GET['quarter']
                    months_in_current_quarter = netperd_in_quarter(int(current_q))
                    net_perds = [month for month in months_in_current_quarter]

                    budget_totals = BudgetTotals.objects.filter(Q(posted=False) & Q(department_id__in=(9, 14,))
                                                                ).order_by('-department_id',
                                                                           '-account_id').values(
                        *net_perds, 'account_id', 'account__acctdesc', 'account_acctfmttd')

                    acctIds = [obj['account_id'] for obj in
                               budget_totals.filter(department_id=department)]
                    acctIdsIncomeDC = [obj['account_id'] for obj in
                                       budget_totals.filter(department_id=14)]

                    queryset = budget_totals.filter(department_id=department)
                    actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIds, fscsdsg='A', fscscurn='ZMW',
                                                                       curntype='F',
                                                                       fscsyr=year.year).values(
                        *net_perds,
                        'acctid')
                    incomedc_queryset = budget_totals.filter(department_id=14)
                    incomedc_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsIncomeDC,
                                                                                fscsyr=year.year).values(
                        *net_perds, 'acctid')

                    querysets = [
                        (f'{request.user.department.name}', department, queryset, actuals),
                        ('Staff Remuneration & Related costs', 14, incomedc_queryset, incomedc_actuals),

                    ]
                    unit_totals = []

                    for dept_name, dept_id, queryset, queryset2 in querysets:
                        actuals = Glafs.objects.using('sql_server').filter(
                            acctid__in=[obj['account_id'] for obj in queryset], fscsdsg='A', fscscurn='ZMW',
                            curntype='F', fscsyr=year.year)
                        values = quarter_aggregate(queryset, budget_totals_all.filter(department_id=dept_id),
                                                   queryset2,
                                                   actuals, current_q, dept_id)
                        unit_totals.append(
                            {
                                'acctdesc': dept_name,
                                **values,
                                'id': dept_id
                            }
                        )
                    data = {'dept_totals': unit_totals}
                    opex_total = opex_total_actual = opex_available = opex_pending = opex_ytd_total = opex_ytd_total_actual = opex_ytd_available = opex_ytd_pending = Decimal(
                        '0.00')

                    for item in data['dept_totals']:
                        opex_total += Decimal(item['total'].replace(',', ''))
                        opex_total_actual += Decimal(item['total_actual'].replace(',', ''))
                        opex_available += Decimal(item['available'].replace(',', ''))
                        opex_pending += Decimal(item['pending'].replace(',', ''))
                        opex_ytd_total += Decimal(item['year_total'].replace(',', ''))
                        opex_ytd_total_actual += Decimal(item['year_total_actual'].replace(',', ''))
                        opex_ytd_available += Decimal(item['available_year_total'].replace(',', ''))
                        opex_ytd_pending += Decimal(item['pending_year_total'].replace(',', ''))

                    context = {'data': data, 'period': f'Q{current_q}', 'department': request.user.department.name,

                               'opex_total': department_total_sum,
                               'opex_actual_total': round(dept_netper_sum, 2),
                               'opex_period_total': opex_total,
                               'opex_period_actual_total': opex_total_actual,
                               'opex_period_pending_total': opex_pending,
                               'opex_period_available_total': opex_available,
                               'opex_ytd_total': opex_ytd_total,
                               'opex_ytd_actual_total': opex_ytd_total_actual,
                               'opex_ytd_pending_total': opex_ytd_pending,
                               'opex_ytd_available_total': opex_ytd_available,

                               }
                    return render(request, 'dept-user-index.html', context)



                else:
                    current_q = current_period()
                    if 'period' in request.GET:
                        current_q = request.GET['period']
                    budget_total = BudgetTotals.objects.filter(department_id__in=(9, 14,), posted=False, ).values(
                        current_q, 'account_id', 'account__acctdesc', 'account_acctfmttd').order_by(
                        '-department_id', '-account_id')
                    budget_total_all = BudgetTotals.objects.filter(department_id__in=(9, 14,),
                                                                   posted=False).order_by(
                        '-department_id', '-account_id')

                    acctIds = [obj['account_id'] for obj in budget_total.filter(department_id=9)]
                    acctIdsIncomeTowers = [obj['account_id'] for obj in
                                           budget_total.filter(department_id=14)]

                    queryset = budget_total.filter(department_id=department)
                    actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIds, fscsdsg='A', fscscurn='ZMW',
                                                                       curntype='F',
                                                                       fscsyr=year.year).values(
                        current_q,
                        'acctid')
                    incometw_queryset = budget_total.filter(department_id=14)
                    incometw_actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIdsIncomeTowers,
                                                                                fscsdsg='A', fscscurn='ZMW',
                                                                                curntype='F',
                                                                                fscsyr=year.year).values(
                        current_q, 'acctid')

                    querysets = [

                        ('Human Capital', 9, queryset, actuals),

                        ('Staff Remuneration & Related costs', 14, incometw_queryset, incometw_actuals),

                    ]

                    unit_totals = []
                    for dept_name, dept_id, queryset, queryset2 in querysets:
                        actuals = Glafs.objects.using('sql_server').filter(
                            acctid__in=[obj['account_id'] for obj in queryset], fscsdsg='A', fscscurn='ZMW',
                            curntype='F', fscsyr=year.year)
                        values = netperd_aggregate(queryset, budget_totals_all.filter(department_id=dept_id), queryset2,
                                                   actuals, current_q, dept_id)
                        unit_totals.append(
                            {
                                'acctdesc': dept_name,
                                **values,
                                'id': dept_id
                            }
                        )

                    data = {'dept_totals': unit_totals}
                    ranges = {
                        'netperd1': 'January',
                        'netperd2': 'February',
                        'netperd3': 'March',
                        'netperd4': 'April',
                        'netperd5': 'May',
                        'netperd6': 'June',
                        'netperd7': 'July',
                        'netperd8': 'August',
                        'netperd9': 'September',
                        'netperd10': 'October',
                        'netperd11': 'November',
                        'netperd12': 'December'
                    }
                    opex_total = opex_total_actual = opex_available = opex_pending = opex_ytd_total = opex_ytd_total_actual = opex_ytd_available = opex_ytd_pending = Decimal(
                        '0.00')

                    for item in data['dept_totals']:
                        if item['id']:
                            opex_total += Decimal(item['total'].replace(',', ''))
                            opex_total_actual += Decimal(item['total_actual'].replace(',', ''))
                            opex_available += Decimal(item['available'].replace(',', ''))
                            opex_pending += Decimal(item['pending'].replace(',', ''))
                            opex_ytd_total += Decimal(item['year_total'].replace(',', ''))
                            opex_ytd_total_actual += Decimal(item['year_total_actual'].replace(',', ''))
                            opex_ytd_available += Decimal(item['available_year_total'].replace(',', ''))
                            opex_ytd_pending += Decimal(item['pending_year_total'].replace(',', ''))

                    context = {'data': data, 'period': ranges[current_q], 'department': request.user.department.name,

                               'opex_total': department_total_sum,
                               'opex_actual_total': round(dept_netper_sum, 2),
                               'opex_period_total': opex_total,
                               'opex_period_actual_total': opex_total_actual,
                               'opex_period_pending_total': opex_pending,
                               'opex_period_available_total': opex_available,
                               'opex_ytd_total': opex_ytd_total,
                               'opex_ytd_actual_total': opex_ytd_total_actual,
                               'opex_ytd_pending_total': opex_ytd_pending,
                               'opex_ytd_available_total': opex_ytd_available,

                               }
                    return render(request, 'dept-user-index.html', context)
            else:
                budget_totals_all = BudgetTotals.objects.filter(
                    Q(posted=False) & Q(department_id=department) & Q(year=year.year) & Q(
                        budget_set=budget_set)).order_by('-department_id',
                                                         '-account_id')

                department_total_sum = budget_totals_all.aggregate(total_sum=Sum('total'))['total_sum'] or 0

                department_actual_total = Glafs.objects.using('sql_server').filter(
                    acctid__in=[obj.account.acctid for obj in budget_totals_all])

                for total in department_actual_total:
                    dept_netper_sum += total.get_netper_sum()

                if 'quarter' in request.GET:
                    current_q = request.GET['quarter']
                    months_in_current_quarter = netperd_in_quarter(int(current_q))
                    net_perds = [month for month in months_in_current_quarter]

                    budget_totals = BudgetTotals.objects.filter(
                        Q(posted=False) & Q(department_id=department) & Q(year=year.year) & Q(budget_set=budget_set)
                        ).order_by('-department_id',
                                   '-account_id').values(
                        *net_perds, 'account_id', 'account__acctdesc', 'account_acctfmttd')

                    acctIds = [obj['account_id'] for obj in
                               budget_totals.filter(department_id=department)]

                    queryset = budget_totals.filter(department_id=department)
                    actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIds, fscsdsg='A', fscscurn='ZMW',
                                                                       curntype='F',
                                                                       fscsyr=year.year).values(
                        *net_perds,
                        'acctid')

                    querysets = [
                        (f'{request.user.department.name}', department, queryset, actuals),
                    ]
                    unit_totals = []

                    for dept_name, dept_id, queryset, queryset2 in querysets:
                        actuals = Glafs.objects.using('sql_server').filter(
                            acctid__in=[obj['account_id'] for obj in queryset], fscsdsg='A', fscscurn='ZMW',
                            curntype='F', fscsyr=year.year)
                        values = quarter_aggregate(queryset, budget_totals_all.filter(department_id=dept_id),
                                                   queryset2,
                                                   actuals, current_q, dept_id)
                        unit_totals.append(
                            {
                                'acctdesc': dept_name,
                                **values,
                                'id': dept_id
                            }
                        )
                    data = {'dept_totals': unit_totals}
                    opex_total = opex_total_actual = opex_available = opex_pending = opex_ytd_total = opex_ytd_total_actual = opex_ytd_available = opex_ytd_pending = Decimal(
                        '0.00')

                    for item in data['dept_totals']:
                        opex_total += Decimal(item['total'].replace(',', ''))
                        opex_total_actual += Decimal(item['total_actual'].replace(',', ''))
                        opex_available += Decimal(item['available'].replace(',', ''))
                        opex_pending += Decimal(item['pending'].replace(',', ''))
                        opex_ytd_total += Decimal(item['year_total'].replace(',', ''))
                        opex_ytd_total_actual += Decimal(item['year_total_actual'].replace(',', ''))
                        opex_ytd_available += Decimal(item['available_year_total'].replace(',', ''))
                        opex_ytd_pending += Decimal(item['pending_year_total'].replace(',', ''))

                    context = {'data': data, 'period': f'Q{current_q}', 'department': request.user.department.name,

                               'opex_total': department_total_sum,
                               'opex_actual_total': round(dept_netper_sum, 2),
                               'opex_period_total': opex_total,
                               'opex_period_actual_total': opex_total_actual,
                               'opex_period_pending_total': opex_pending,
                               'opex_period_available_total': opex_available,
                               'opex_ytd_total': opex_ytd_total,
                               'opex_ytd_actual_total': opex_ytd_total_actual,
                               'opex_ytd_pending_total': opex_ytd_pending,
                               'opex_ytd_available_total': opex_ytd_available,

                               }
                    return render(request, 'dept-user-index.html', context)



                else:
                    current_q = current_period()
                    if 'period' in request.GET:
                        current_q = request.GET['period']
                    budget_total = BudgetTotals.objects.filter(department_id=department, posted=False).values(current_q,
                                                                                                              'account_id',
                                                                                                              'account__acctdesc',
                                                                                                              'account__acctfmttd').order_by(
                        '-department_id', '-account_id')
                    budget_total_all = BudgetTotals.objects.filter(department_id=department, posted=False).exclude(
                        total=0).order_by(
                        '-department_id', '-account_id')

                    acctIds = [obj['account_id'] for obj in budget_total.filter(department_id=department)]

                    queryset = budget_total.filter(department_id=department)
                    actuals = Glafs.objects.using('sql_server').filter(acctid__in=acctIds, fscsdsg='A', fscscurn='ZMW',
                                                                       curntype='F',
                                                                       fscsyr=year.year).values(
                        current_q,
                        'acctid')
                    unit_totals = []
                    querysets = [
                        (f'{request.user.department.name}', department, queryset, actuals),
                    ]
                    for dept_name, dept_id, queryset, queryset2 in querysets:
                        actuals = Glafs.objects.using('sql_server').filter(
                            acctid__in=[obj['account_id'] for obj in queryset], fscsdsg='A', fscscurn='ZMW',
                            curntype='F', fscsyr=year.year)
                        values = netperd_aggregate(queryset, budget_totals_all.filter(department_id=dept_id), queryset2,
                                                   actuals, current_q, dept_id)
                        unit_totals.append(
                            {
                                'acctdesc': dept_name,
                                **values,
                                'id': dept_id
                            }
                        )

                    data = {'dept_totals': unit_totals}
                    ranges = {
                        'netperd1': 'January',
                        'netperd2': 'February',
                        'netperd3': 'March',
                        'netperd4': 'April',
                        'netperd5': 'May',
                        'netperd6': 'June',
                        'netperd7': 'July',
                        'netperd8': 'August',
                        'netperd9': 'September',
                        'netperd10': 'October',
                        'netperd11': 'November',
                        'netperd12': 'December'
                    }
                    opex_total = opex_total_actual = opex_available = opex_pending = opex_ytd_total = opex_ytd_total_actual = opex_ytd_available = opex_ytd_pending = Decimal(
                        '0.00')

                    for item in data['dept_totals']:
                        opex_total += Decimal(item['total'])
                        opex_total_actual += Decimal(item['total_actual'])
                        opex_available += Decimal(item['available'].replace(',', ''))
                        opex_pending += Decimal(item['pending'])
                        opex_ytd_total += Decimal(item['year_total'])
                        opex_ytd_total_actual += Decimal(item['year_total_actual'])
                        opex_ytd_available += Decimal(item['available_year_total'].replace(',', ''))
                        opex_ytd_pending += Decimal(item['pending_year_total'])

                    context = {'data': data, 'period': ranges[current_q], 'department': request.user.department.name,

                               'opex_total': department_total_sum,
                               'opex_actual_total': round(dept_netper_sum, 2),
                               'opex_period_total': opex_total,
                               'opex_period_actual_total': opex_total_actual,
                               'opex_period_pending_total': opex_pending,
                               'opex_period_available_total': opex_available,
                               'opex_ytd_total': opex_ytd_total,
                               'opex_ytd_actual_total': opex_ytd_total_actual,
                               'opex_ytd_pending_total': opex_ytd_pending,
                               'opex_ytd_available_total': opex_ytd_available,

                               }
                    return render(request, 'dept-user-index.html', context)

    else:
        return redirect('budgets:login')


def capex_index(request):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)
        obj = BudgetVariations.objects.get(is_active=True,year=year.year)
        account_details = BudgetTotals.objects.filter(budget_set=obj.budget_set, posted=False, year=year.year).order_by(
            '-account_id').annotate(
            count=Count('department__id'))

        # Income accounts

        # Asset accounts
        asset = BudgetTotals.objects.filter(budget_set=obj.budget_set, department_id=13, posted=False,
                                            year=year.year).order_by(
            '-account_id')
        paginator = Paginator(asset, 10)
        page_number = request.GET.get('page')
        asset_obj = paginator.get_page(page_number)

        total_sum = account_details.aggregate(total_sum=Sum('total'))['total_sum']

        context = {
            'asset': asset_obj,
            'budget_set': obj.budget_set,
            'status': obj.is_complete,
            'total': total_sum
        }
        return render(request, 'capex-index.html', context)
    else:
        return redirect('budgets:login')


def dept_user_index(request):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)

        obj = BudgetVariations.objects.get(is_active=True,year=year.year)
        department_conditions = Q()
        for department_id in range(1, 13):  # Assuming departments IDs range from 1 to 12
            department_conditions |= Q(department_id=department_id)
        active = get_object_or_404(BudgetVariations, is_active=True, year=year.year)
        # Fetch all relevant budget totals in a single query
        budget_totals = BudgetTotals.objects.filter(
            Q(budget_set=obj.budget_set) & Q(posted=False) & Q(year=year.year) & department_conditions
        ).order_by('-department_id', '-account_id')

        # Create separate variables for each department
        ceo = budget_totals.filter(department_id=1)
        paginator = Paginator(ceo, 10)
        page_number = request.GET.get('page')
        ceo_obj = paginator.get_page(page_number)
        internal_audit = budget_totals.filter(department_id=2)
        paginator = Paginator(internal_audit, 10)
        page_number = request.GET.get('page')
        internal_audit_obj = paginator.get_page(page_number)
        supply_chain = budget_totals.filter(department_id=3)
        paginator = Paginator(supply_chain, 10)
        page_number = request.GET.get('page')
        supply_chain_obj = paginator.get_page(page_number)
        bds = budget_totals.filter(department_id=4)
        paginator = Paginator(bds, 10)
        page_number = request.GET.get('page')
        bds_obj = paginator.get_page(page_number)
        public_relations = budget_totals.filter(department_id=10)
        paginator = Paginator(public_relations, 10)
        page_number = request.GET.get('page')
        public_relations_obj = paginator.get_page(page_number)
        technical = budget_totals.filter(department_id=6)
        paginator = Paginator(technical, 10)
        page_number = request.GET.get('page')
        technical_obj = paginator.get_page(page_number)
        information_systems = budget_totals.filter(department_id=7)
        paginator = Paginator(information_systems, 10)
        page_number = request.GET.get('page')
        information_systems_obj = paginator.get_page(page_number)
        legal_risk = budget_totals.filter(department_id=8)
        paginator = Paginator(legal_risk, 10)
        page_number = request.GET.get('page')
        legal_risk_obj = paginator.get_page(page_number)
        human_capital = budget_totals.filter(department_id=9)
        paginator = Paginator(human_capital, 10)
        page_number = request.GET.get('page')
        human_capital_obj = paginator.get_page(page_number)
        sales_marketing = budget_totals.filter(department_id=10)
        paginator = Paginator(sales_marketing, 10)
        page_number = request.GET.get('page')
        sales_marketing_obj = paginator.get_page(page_number)
        admin = budget_totals.filter(department_id=11)
        paginator = Paginator(admin, 10)
        page_number = request.GET.get('page')
        admin_page_obj = paginator.get_page(page_number)
        finance = budget_totals.filter(department_id=12)
        paginator = Paginator(finance, 10)
        page_number = request.GET.get('page')
        finance_page_obj = paginator.get_page(page_number)
        # Calculate total sum
        total_sum = budget_totals.filter(department=request.user.department).aggregate(total_sum=Sum('total'))[
            'total_sum']

        # Populate context variable
        context = {
            'ceo': ceo_obj,
            'internal_audit': internal_audit_obj,
            'supply_chain': supply_chain_obj,
            'bds': bds_obj,
            'public_relations': public_relations_obj,
            'technical': technical_obj,
            'information_systems': information_systems_obj,
            'legal_risk': legal_risk_obj,
            'human_capital': human_capital_obj,
            'sales_marketing': sales_marketing_obj,
            'admin': admin_page_obj,
            'finance': finance_page_obj,
            'budget_set': obj.budget_set,
            'total': total_sum,
            'activebs': active.budget_set
        }
        return render(request, 'index-deptuser.html', context)
    else:
        return redirect('budgets:login')

def get_paginated_data(budget_totals, department_id, per_page=10):
            """Helper function to filter, paginate, and extract group IDs for a department."""
            queryset = budget_totals.filter(department_id=department_id)

            group_ids = Group.objects.filter(department_id=department_id, parent_group__isnull=True).order_by('name')

            # Include sub-groups within each parent group
            groups_with_subgroups = []
            for group in group_ids:
                sub_groups = Group.objects.filter(parent_group=group)
                groups_with_subgroups.append((group, sub_groups))

            return queryset, groups_with_subgroups


def budget_summary(request):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)
        obj = BudgetVariations.objects.get(is_active=True,year=year.year)
        if request.user:

            # Combine all department conditions using Q objects
            department_conditions = Q()
            for department_id in range(1, 15):  # Assuming departments IDs range from 1 to 12
                department_conditions |= Q(department_id=department_id)

            # Fetch all relevant budget totals in a single query
            budget_totals = BudgetLines.objects.filter(
                Q(budget_set=obj.budget_set) & Q(year=year.year) & Q(posted=False) & department_conditions
            ).exclude(department_id=13).exclude(total=0).order_by('-department_id', '-account_id')
            groups = Group.objects.all()

            # Your main view logic
            ceo_obj, ceo_group_ids = get_paginated_data(budget_totals, 1)
            internal_audit_obj, internal_audit_group_ids = get_paginated_data(budget_totals, 2)
            supply_chain_obj, supply_chain_group_ids = get_paginated_data(budget_totals, 3)
            bds_obj, bds_group_ids = get_paginated_data(budget_totals, 4)
            public_relations_obj, public_relations_group_ids = get_paginated_data(budget_totals, 10)
            technical_obj, technical_group_ids = get_paginated_data(budget_totals, 6)
            information_systems_obj, information_systems_group_ids = get_paginated_data(budget_totals, 7)
            legal_risk_obj, legal_risk_group_ids = get_paginated_data(budget_totals, 8)
            human_capital_obj, human_capital_group_ids = get_paginated_data(budget_totals, 9)
            sales_marketing_obj, sales_marketing_group_ids = get_paginated_data(budget_totals, 10)
            admin_page_obj, admin_group_ids = get_paginated_data(budget_totals, 11)
            finance_page_obj, finance_group_ids = get_paginated_data(budget_totals, 12)
            staff_n_page_obj, staff_n_group_ids = get_paginated_data(budget_totals, 14)
            # Calculate total sum
            total_sum = budget_totals.aggregate(total_sum=Sum('total'))['total_sum']
            # Populate context variable
            context = {
                'status': obj.is_complete,
                'ceo': ceo_obj,
                'ceo_group_ids': ceo_group_ids,
                'internal_audit': internal_audit_obj,
                'internal_audit_group_ids': internal_audit_group_ids,
                'supply_chain': supply_chain_obj,
                'supply_chain_group_ids': supply_chain_group_ids,
                'bds': bds_obj,
                'bds_group_ids': bds_group_ids,
                'public_relations': public_relations_obj,
                'public_relations_group_ids': public_relations_group_ids,
                'technical': technical_obj,
                'technical_group_ids': technical_group_ids,
                'information_systems': information_systems_obj,
                'information_systems_group_ids': information_systems_group_ids,
                'legal_risk': legal_risk_obj,
                'legal_risk_group_ids': legal_risk_group_ids,
                'human_capital': human_capital_obj,
                'human_capital_group_ids': human_capital_group_ids,
                'sales_marketing': sales_marketing_obj,
                'sales_marketing_group_ids': sales_marketing_group_ids,
                'admin': admin_page_obj,
                'admin_group_ids': admin_group_ids,

                'finance': finance_page_obj,
                'finance_group_ids': finance_group_ids,
                'staff_n_page': staff_n_page_obj,
                'staff_n_group_ids': staff_n_group_ids,
                'budget_set': obj.budget_set,
                'total': total_sum
            }

            return render(request, 'budget_summary.html', context)



    else:
        return redirect('budgets:login')

def replicate_accounts_with_group_ids(request):
    # Define the account values to filter
    values = [
        "62040-200", "62070-200", "62080-200",
        "62100-200", "62110-200", "62190-200",
        "62195-200", "62225-200", "63270-200"
    ]

    # Hardcoded list of group IDs to assign
    group_ids = [170, 172, 173, 174, 175, 176]

    # Retrieve the accounts based on the specified values
    query = Q()
    for value in values:
        query |= Q(acctfmttd__icontains=value)

    # Retrieve the accounts based on the specified values using icontains
    accounts = Accounts.objects.filter(query)
    new_accounts = []
    # Loop through existing accounts and replicate them
    for account in accounts:
        for group_id in group_ids:
            # Create a new account instance with the same data but different group_id
            new_account = Accounts(
                acctid=account.acctid,  # Assuming acctid can be duplicated
                group_id=group_id,  # Set the new group_id
                audtdate=account.audtdate,
                audttime=account.audttime,
                audtuser=account.audtuser,
                audtorg=account.audtorg,
                createdate=account.createdate,
                acctdesc=account.acctdesc,
                accttype=account.accttype,
                acctbal=account.acctbal,
                activesw=account.activesw,
                consldsw=account.consldsw,
                qtysw=account.qtysw,
                uom=account.uom,
                allocsw=account.allocsw,
                acctofset=account.acctofset,
                acctsrty=account.acctsrty,
                mcsw=account.mcsw,
                specsw=account.specsw,
                acctgrpcod=account.acctgrpcod,
                ctrlacctsw=account.ctrlacctsw,
                srceldgid=account.srceldgid,
                alloctot=account.alloctot,
                abrkid=account.abrkid,
                yracctclos=account.yracctclos,
                acctfmttd=account.acctfmttd,
                acsegval01=account.acsegval01,
                acsegval02=account.acsegval02,
                acsegval03=account.acsegval03,
                acsegval04=account.acsegval04,
                acsegval05=account.acsegval05,
                acsegval06=account.acsegval06,
                acsegval07=account.acsegval07,
                acsegval08=account.acsegval08,
                acsegval09=account.acsegval09,
                acsegval10=account.acsegval10,
                acctsegval=account.acctsegval,
                acctgrpscd=account.acctgrpscd,
                postosegid=account.postosegid,
                defcurncod=account.defcurncod,
                ovalues=account.ovalues,
                tovalues=account.tovalues,
                rollupsw=account.rollupsw,
                department=account.department)
            new_accounts.append(new_account)

    # Bulk create new account instances
    Accounts.objects.bulk_create(new_accounts)

    # Return a JSON response indicating success
    return JsonResponse({'status': 'success', 'created_count': len(new_accounts)})
def load_groups(request):
    departments = Department.objects.all()
    values  = ["71492-800", "71494-800", "71500-800"]

    groups = Group.objects.filter(name='Staff Welfare')
    dept_to_group = {group.department_id: group.id for group in groups}
    for acc in values:
        account = Accounts.objects.filter(acctfmttd__icontains=acc)
        if account.exists():
            for account in account:
                group_id = dept_to_group.get(account.department_id)
                print(account, group_id)
                Accounts.objects.filter(acctid=account.acctid).update(group_id=group_id)
        else:
            continue
"""from openpyxl.utils.dataframe import dataframe_to_rows
def to_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=budget_summary.xlsx'
    data = {
        'Department': ['Sales', 'Sales', 'Sales', 'HR', 'HR', 'HR', 'Finance', 'Finance', 'IT', 'IT'],
        'Group': ['Team A', 'Team B', 'Team C', 'Team D', 'Team E', 'Team F', 'Team G', 'Team H', 'Team I', 'Team J'],
        'Employee': ['Alice', 'Bob', 'Charlie', 'David', 'Eva', 'Frank', 'Grace', 'Hank', 'Ivy', 'Jack'],
        'Salary': [50000, 60000, 55000, 52000, 58000, 57000, 62000, 61000, 65000, 68000]
    }

    # Create a DataFrame
    df = DataFrame(data)

    # Sort the data by Department, Group, and Employee
    df_sorted = df.sort_values(by=['Department', 'Group', 'Employee'])

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ['Department', 'Group', 'Employee', 'Salary']
    ws.append(headers)

    # Define the starting row index
    start_row = 2

    # Add data to the Excel sheet with hierarchical grouping
    for dept in df_sorted['Department'].unique():
        dept_data = df_sorted[df_sorted['Department'] == dept]
        dept_start_row = start_row

        for group in dept_data['Group'].unique():
            group_data = dept_data[dept_data['Group'] == group]
            group_start_row = start_row

            for row in dataframe_to_rows(group_data, index=False, header=False):
                ws.append(row)
                start_row += 1

            # Group the rows for each group (Level 3 - Employee level)
            ws.row_dimensions.group(group_start_row, start_row - 1, hidden=True)

        # Group the rows for each department (Level 2 - Group level)
        ws.row_dimensions.group(dept_start_row, start_row - 1, hidden=False)

        # Department-level grouping (Level 1)
        ws.row_dimensions.group(dept_start_row, start_row - 1, hidden=False)
    wb.save(response)
    return response"""

def to_excel_department(request,dept_id):
    output = io.BytesIO()
    dept = Department.objects.get(id=dept_id)
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({"bold": True})
    center_format = workbook.add_format({"bold": True, "align": "center", "valign": "vcenter", "font_size": 14})
    worksheet.merge_range('A1:R1', f'Budget Summary {dept.name}', center_format)  # Adjust the range as needed
    worksheet.set_row(0, 30)

    # Fetch budget data
    year = FinancialYear.objects.get(is_active=True)
    obj = BudgetVariations.objects.get(is_active=True,year=year.year)

    # Fetch budget totals
    budget_totals = BudgetLines.objects.filter(
        Q(budget_set=obj.budget_set) & Q(year=year.year) & Q(department_id=dept_id)
    ).exclude(total=0).order_by('department_id', 'group_id')

    # Initialize variables for tracking
    current_department = None
    current_group = None
    row = 4  # Start writing data from the first row

    # Iterate through budget totals and write to the worksheet
    for line in budget_totals:
        # Check if we need to start a new department group
        if line.department_id != current_department:
            if current_department is not None:
                row += 1  # Move to the next row for the new department

            current_department = line.department_id
            worksheet.write(row, 0, line.department.name,bold)  # Write department name
            worksheet.set_row(row, None, None, {"level": 1})  # Set outline level for department
            row += 1  # Move to the next row

        # Check if we need to start a new group
        if line.account.group_id != current_group:
            current_group = line.account.group_id
            # Check if line.group is None before accessing its name
            group_name = line.account.group.name if line.account.group else ""
            worksheet.write(row, 1, group_name,bold)  # Write group name
            worksheet.set_row(row, None, None, {"level": 2})  # Set outline level for group
            row += 1  # Move to the next row

            # Write headers for the current group
            headers = [
                "Account ID", "Name", "Currency", "Year",
                "Total", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
            ]
            worksheet.write_row(row, 2, headers, bold)  # Write headers starting from column C
            worksheet.set_row(row, None, None, {"level": 3})  # Set outline level for headers
            row += 1  # Move to the next row after headers

        # Write the budget line details
        worksheet.write(row, 2, line.account_id)  # Write Account ID
        worksheet.write(row, 3, line.item_description)  # Write item description
        worksheet.write(row, 4, line.currency.currency)  # Write currency
        worksheet.write(row, 5, line.year)  # Write year
        worksheet.write(row, 6, line.total)  # Write total
        worksheet.write(row, 7, line.netperd1)  # Write January value
        worksheet.write(row, 8, line.netperd2)  # Write February value
        worksheet.write(row, 9, line.netperd3)  # Write March value
        worksheet.write(row, 10, line.netperd4)  # Write April value
        worksheet.write(row, 11, line.netperd5)  # Write May value
        worksheet.write(row, 12, line.netperd6)  # Write June value
        worksheet.write(row, 13, line.netperd7)  # Write July value
        worksheet.write(row, 14, line.netperd8)  # Write August value
        worksheet.write(row, 15, line.netperd9)  # Write September value
        worksheet.write(row, 16, line.netperd10)  # Write October value
        worksheet.write(row, 17, line.netperd11)  # Write November value
        worksheet.write(row, 18, line.netperd12)  # Write December value
        worksheet.set_row(row, None, None, {"level": 3})  # Set outline level for data rows
        row += 1  # Move to the next row

    # Close the workbook
    workbook.close()

    # Set up the response
    output.seek(0)
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="budget_summary.xlsx"'
    return response
def to_excel(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet()
    bold = workbook.add_format({"bold": True})
    center_format = workbook.add_format({"bold": True, "align": "center", "valign": "vcenter", "font_size": 14})
    worksheet.merge_range('A1:R1', 'Budget Summary', center_format)  # Adjust the range as needed
    worksheet.set_row(0, 30)

    # Fetch budget data
    year = FinancialYear.objects.get(is_active=True)
    obj = BudgetVariations.objects.get(is_active=True,year=year.year)

    # Fetch budget totals
    budget_totals = BudgetLines.objects.filter(
        Q(budget_set=obj.budget_set) & Q(year=year.year) & Q(posted=False)
    ).exclude(department_id=13).exclude(total=0).order_by('department_id', 'group_id')

    # Initialize variables for tracking
    current_department = None
    current_group = None
    row = 4  # Start writing data from the first row

    # Iterate through budget totals and write to the worksheet
    for line in budget_totals:
        # Check if we need to start a new department group
        if line.department_id != current_department:
            if current_department is not None:
                row += 1  # Move to the next row for the new department

            current_department = line.department_id
            worksheet.write(row, 0, line.department.name,bold)  # Write department name
            worksheet.set_row(row, None, None, {"level": 1})  # Set outline level for department
            row += 1  # Move to the next row

        # Check if we need to start a new group
        if line.account.group_id != current_group:
            current_group = line.account.group_id
            # Check if line.group is None before accessing its name
            group_name = line.account.group.name if line.account.group else ""
            worksheet.write(row, 1, group_name,bold)  # Write group name
            worksheet.set_row(row, None, None, {"level": 2})  # Set outline level for group
            row += 1  # Move to the next row

            # Write headers for the current group
            headers = [
                "Account ID", "Name", "Currency", "Year",
                "Total", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
            ]
            worksheet.write_row(row, 2, headers, bold)  # Write headers starting from column C
            worksheet.set_row(row, None, None, {"level": 3})  # Set outline level for headers
            row += 1  # Move to the next row after headers

        # Write the budget line details
        worksheet.write(row, 2, line.account_id)  # Write Account ID
        worksheet.write(row, 3, line.item_description)  # Write item description
        worksheet.write(row, 4, line.currency.currency)  # Write currency
        worksheet.write(row, 5, line.year)  # Write year
        worksheet.write(row, 6, line.total)  # Write total
        worksheet.write(row, 7, line.netperd1)  # Write January value
        worksheet.write(row, 8, line.netperd2)  # Write February value
        worksheet.write(row, 9, line.netperd3)  # Write March value
        worksheet.write(row, 10, line.netperd4)  # Write April value
        worksheet.write(row, 11, line.netperd5)  # Write May value
        worksheet.write(row, 12, line.netperd6)  # Write June value
        worksheet.write(row, 13, line.netperd7)  # Write July value
        worksheet.write(row, 14, line.netperd8)  # Write August value
        worksheet.write(row, 15, line.netperd9)  # Write September value
        worksheet.write(row, 16, line.netperd10)  # Write October value
        worksheet.write(row, 17, line.netperd11)  # Write November value
        worksheet.write(row, 18, line.netperd12)  # Write December value
        worksheet.set_row(row, None, None, {"level": 3})  # Set outline level for data rows
        row += 1  # Move to the next row

    # Close the workbook
    workbook.close()

    # Set up the response
    output.seek(0)
    response = HttpResponse(output.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="budget_summary.xlsx"'
    return response
def account_list(request):
    # Retrieve all accounts and annotate them by their group
    accounts = Accounts.objects.select_related('group').all()

    # Group accounts by their current group
    grouped_accounts = {}
    for account in accounts:
        group_name = account.group.name if account.group else 'Unassigned'
        if group_name not in grouped_accounts:
            grouped_accounts[group_name] = []
        grouped_accounts[group_name].append(account)
    print(grouped_accounts)
    return render(request, 'account_list.html', {'grouped_accounts': grouped_accounts })
def group_index_dept(request,id):
    if request.user.is_authenticated:

            dept = Department.objects.filter(id=id)
            depart = Department.objects.get(id=id)
            accounts_without_group = Accounts.objects.filter(group__isnull=False,
                                                                 department_id=id).select_related(
                    'group')

                # Group accounts by their current group (in this case, it will only be 'Unassigned')
            grouped_accounts = {f'{depart.name}': list(accounts_without_group)}

            return render(request, 'account-group.html', {'grouped_accounts': grouped_accounts, 'department': dept})
def group_index(request):
    if request.user.is_authenticated:
        if request.GET.get('dept'):
            dept = Department.objects.filter(id=request.GET.get('dept'))
            depart = Department.objects.get(id=request.GET.get('dept'))
            accounts_without_group = Accounts.objects.filter(group__isnull=False,
                                                                 department_id=request.GET.get('dept')).select_related(
                    'group')

                # Group accounts by their current group (in this case, it will only be 'Unassigned')
            grouped_accounts = {f'{depart.name}': list(accounts_without_group)}

            return render(request, 'account-group.html', {'grouped_accounts': grouped_accounts, 'department': dept})
        else:

            if request.user.role == '002':
                accounts_without_group = Accounts.objects.filter(group__isnull=True).exclude(department_id__in=[17,13]).select_related('group')

                # Group accounts by their current group (in this case, it will only be 'Unassigned')
                grouped_accounts = {'Unassigned': list(accounts_without_group)}
                dept = Department.objects.exclude(id__in=[17,13])

                # Fetch all departments
                print('admin')
                return render(request, 'account-group.html', {'grouped_accounts': grouped_accounts, 'department': dept})


            else:
                accounts_without_group = Accounts.objects.filter(group__isnull=True,department_id= request.user.department.id).select_related('group')

                # Group accounts by their current group (in this case, it will only be 'Unassigned')
                grouped_accounts = {'Unassigned': list(accounts_without_group)}
                dept = Department.objects.filter(id=request.user.department.id)
                print('ordinary user')
                return render(request, 'account-group.html', {'grouped_accounts': grouped_accounts, 'department': dept})


def account_line_search(request):
    year = FinancialYear.objects.get(is_active=True)
    acc = request.GET.get('acc'
                        )
    lines = BudgetLines.objects.filter(account__acctid__icontains=acc,year=year.year).values('id','item_description')
    if lines:
        return  JsonResponse({'data':list(lines)},status=200)
    else:
        return JsonResponse({'data':{}},status=200)

def vary_budget(request,object_id):
    year = FinancialYear.objects.get(is_active=True)

    line = get_object_or_404(BudgetLines, id=object_id, year=year.year)

    active = get_object_or_404(BudgetVariations, is_active=True, year=year.year)
    obj_t = get_object_or_404(BudgetTotals, account_id=line.account_id, budget_set=active.budget_set, year=year.year)
    print(obj_t)
    if active:
        if request.user.role != '002':
            single_object = BudgetTotals.objects.filter(budget_set=active.budget_set,
                                                        department=request.user.department, year=year.year)
        else:
            single_object = BudgetTotals.objects.filter(budget_set=active.budget_set, year=year.year)


        user = get_object_or_404(Users, id=request.user.id)
       # obj = get_object_or_404(BudgetTotals, id=first_dept_obj.id, year=year.year)

        context = {
                'currency': Currency.objects.all(),
                'name': line.account.acctdesc,
                'obj': single_object,
                'object': obj_t,
                'activebs': active.budget_set,
                'department': obj_t.department.name,
                'line': line,
                'budget_set': active.budget_set,
            }

    return render(request,"variation.html", context)
def assign_group(request):
    account = request.GET.get('account')
    group = request.GET.get('group')
    # Validate that both account and group are provided
    if not account or not group:
        return JsonResponse({'message': 'Account and group parameters are required.'}, status=400)

    # Use get_object_or_404 to handle account retrieval
    acc = get_object_or_404(Accounts, acctfmttd=account)

    # Assign the group to the account
    acc.group_id = group
    acc.save()

    return JsonResponse({'message': 'Successful'}, status=200)

def department_group(request):
    dept = request.GET.get('dept')
    group = Group.objects.filter(department_id=dept).values('id','name')
    return JsonResponse({'data': list(group) }, status=200)

def opex_index(request):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)
        obj = BudgetVariations.objects.get(is_active=True,year=year.year)

        # Combine all department conditions using Q objects
        department_conditions = Q()
        for department_id in range(1, 15):  # Assuming departments IDs range from 1 to 12
            department_conditions |= Q(department_id=department_id)

        # Fetch all relevant budget totals in a single query
        budget_totals = BudgetTotals.objects.filter(
            Q(budget_set=obj.budget_set) & Q(year=year.year) & Q(posted=False) & department_conditions
        ).exclude(department_id=13).order_by('-department_id', '-account_id')
        groups = Group.objects.all()

        def get_paginated_data(budget_totals, department_id, per_page=10):
            """Helper function to filter, paginate, and extract group IDs for a department."""
            queryset = budget_totals.filter(department_id=department_id)
            paginator = Paginator(queryset,per_page )
            page_number = request.GET.get('page')
            queryset_obj = paginator.get_page(page_number)
            group_ids = Group.objects.filter(department_id=department_id, parent_group__isnull=True).order_by('name')

            # Include sub-groups within each parent group
            groups_with_subgroups = []
            for group in group_ids:
                sub_groups = Group.objects.filter(parent_group=group)
                groups_with_subgroups.append((group, sub_groups))

            return queryset_obj, groups_with_subgroups

        # Your main view logic
        ceo_obj, ceo_group_ids = get_paginated_data(budget_totals, 1)
        internal_audit_obj, internal_audit_group_ids = get_paginated_data(budget_totals, 2)
        supply_chain_obj, supply_chain_group_ids = get_paginated_data(budget_totals, 3)
        bds_obj, bds_group_ids = get_paginated_data(budget_totals, 4)
        public_relations_obj, public_relations_group_ids = get_paginated_data(budget_totals, 10)
        technical_obj, technical_group_ids = get_paginated_data(budget_totals, 6)
        information_systems_obj, information_systems_group_ids = get_paginated_data(budget_totals, 7)
        legal_risk_obj, legal_risk_group_ids = get_paginated_data(budget_totals, 8)
        human_capital_obj, human_capital_group_ids = get_paginated_data(budget_totals, 9)
        sales_marketing_obj, sales_marketing_group_ids = get_paginated_data(budget_totals, 10)
        admin_page_obj, admin_group_ids = get_paginated_data(budget_totals, 11)
        finance_page_obj, finance_group_ids = get_paginated_data(budget_totals, 12)
        staff_n_page_obj, staff_n_group_ids = get_paginated_data(budget_totals, 14)
        # Calculate total sum
        total_sum = budget_totals.aggregate(total_sum=Sum('total'))['total_sum']
        # Populate context variable
        context = {
            'status': obj.is_complete,
            'ceo': ceo_obj,
            'ceo_group_ids': ceo_group_ids,
            'internal_audit': internal_audit_obj,
            'internal_audit_group_ids': internal_audit_group_ids,
            'supply_chain': supply_chain_obj,
            'supply_chain_group_ids': supply_chain_group_ids,
            'bds': bds_obj,
            'bds_group_ids': bds_group_ids,
            'public_relations': public_relations_obj,
            'public_relations_group_ids': public_relations_group_ids,
            'technical': technical_obj,
            'technical_group_ids': technical_group_ids,
            'information_systems': information_systems_obj,
            'information_systems_group_ids': information_systems_group_ids,
            'legal_risk': legal_risk_obj,
            'legal_risk_group_ids': legal_risk_group_ids,
            'human_capital': human_capital_obj,
            'human_capital_group_ids': human_capital_group_ids,
            'sales_marketing': sales_marketing_obj,
            'sales_marketing_group_ids': sales_marketing_group_ids,
            'admin': admin_page_obj,
            'admin_group_ids': admin_group_ids,

            'finance': finance_page_obj,
            'finance_group_ids': finance_group_ids,
            'staff_remuneration': staff_n_page_obj,
            'staff_remuneration_group_ids': staff_n_group_ids,
            'budget_set': obj.budget_set,
            'total': total_sum
        }
        return render(request, 'opex-index.html', context)


def reports(request):
    excel_header = [
        'Account Number',
        'Account Name',
        'netperd 1 (Budget)',
        'netperd 1 (Actual)',
        'netperd 2 (Budget)',
        'netperd 2 (Actual)',
        'netperd 3 (Budget)',
        'netperd 3 (Actual)',
        'netperd 4 (Budget)',
        'netperd 4 (Actual)',
        'netperd 5 (Budget)',
        'netperd 5 (Actual)',
        'netperd 6 (Budget)',
        'netperd 6 (Actual)',
        'netperd 7 (Budget)',
        'netperd 7 (Actual)',
        'netperd 8 (Budget)',
        'netperd 8 (Actual)',
        'netperd 9 (Budget)',
        'netperd 9 (Actual)',
        'netperd 10 (Budget)',
        'netperd 10 (Actual)',
        'netperd 11 (Budget)',
        'netperd 11 (Actual)',
        'netperd 12 (Budget)',
        'netperd 12 (Actual)',
        'Q1 (Actual)',
        'Q1 (Budget)',
        'Q2 (Budget)',
        'Q2 (Actual)',
        'Q3 (Budget)',
        'Q3 (Actual)',
        'Q4 (Budget)',
        'Q4 (Actual)'
    ]
    year = FinancialYear.objects.get(is_active=True)

    # Grouping the values according to quarters and halves
    budget_totals = defaultdict(lambda: defaultdict(Decimal))
    budget_object = BudgetTotals.objects.exclude(department_id=13)
    budget_actual_id = [obj.account_id for obj in budget_object]
    budget_actuals = Glafs.objects.using('sql_server').filter(acctid__in=budget_actual_id, fscsdsg='A', fscscurn='ZMW',
                                                              curntype='F',
                                                              fscsyr=year.year)
    for budget_total, actual_total in zip(budget_object, budget_actuals):
        for netperd in range(1, 13):
            value_budget = getattr(budget_total, f'netperd{netperd}', Decimal('0'))
            value_actual = getattr(actual_total, f'netperd{netperd}', Decimal('0'))
            budget_totals[budget_total.account][f'netperd {netperd} (Budget)'] = value_budget
            budget_totals[budget_total.account][f'netperd {netperd} (Actual)'] = value_actual
            budget_totals[budget_total.account]['Account Name'] = budget_total.account.acctdesc
            for quarter in range(1, 5):
                start_netperd = (quarter - 1) * 3 + 1
                end_netperd = quarter * 3
                quarter_value_budget = sum(getattr(budget_total, f'netperd{netperd}', Decimal('0')) for netperd in
                                           range(start_netperd, end_netperd + 1))
                quarter_value_actual = sum(getattr(actual_total, f'netperd{netperd}', Decimal('0')) for netperd in
                                           range(start_netperd, end_netperd + 1))

                budget_totals[budget_total.account][f'Q{quarter} (Budget)'] = quarter_value_budget
                budget_totals[budget_total.account][f'Q{quarter} (Actual)'] = quarter_value_actual
    return render(request, 'opex-reports.html')


def changelog_report(request):
    excel_header = [
        'Account Number',
        'Account Name',
        'Amount',
        'Action',
        'Department',
        'User',

    ]

    # Grouping the values according to quarters and halves
    budget_totals = defaultdict(lambda: defaultdict(Decimal))
    budget_object = ChangeLog.objects.values('user__firstname', 'account_id', 'account__acctdesc', 'amount', 'action',
                                             'department')

    for budget_total in budget_object:
        pass
    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active

    # Writing the header row with bold and blue background color
    for col, header in enumerate(excel_header, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = cell.font.copy(bold=True, color=colors.WHITE)
        cell.fill = cell.fill.copy(fill_type='solid', start_color='00008B')

    # Writing the data rows

    # Create a HttpResponse with an Excel file attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'inline; filename="changelog-export {datetime.today().date()}.xlsx"'

    # Save the workbook to the HttpResponse
    workbook.save(response)

    return response


def generate_excel_opex(request):
    # Define the header row for the Excel file
    excel_header = [
        'Account Number',
        'Account Name',
        'netperd 1 (Budget)',
        'netperd 1 (Actual)',
        'netperd 2 (Budget)',
        'netperd 2 (Actual)',
        'netperd 3 (Budget)',
        'netperd 3 (Actual)',
        'netperd 4 (Budget)',
        'netperd 4 (Actual)',
        'netperd 5 (Budget)',
        'netperd 5 (Actual)',
        'netperd 6 (Budget)',
        'netperd 6 (Actual)',
        'netperd 7 (Budget)',
        'netperd 7 (Actual)',
        'netperd 8 (Budget)',
        'netperd 8 (Actual)',
        'netperd 9 (Budget)',
        'netperd 9 (Actual)',
        'netperd 10 (Budget)',
        'netperd 10 (Actual)',
        'netperd 11 (Budget)',
        'netperd 11 (Actual)',
        'netperd 12 (Budget)',
        'netperd 12 (Actual)',
        'Q1 (Actual)',
        'Q1 (Budget)',
        'Q2 (Budget)',
        'Q2 (Actual)',
        'Q3 (Budget)',
        'Q3 (Actual)',
        'Q4 (Budget)',
        'Q4 (Actual)',
    ]

    year = FinancialYear.objects.get(is_active=True)
    # Grouping the values according to quarters and halves
    budget_totals = defaultdict(lambda: defaultdict(Decimal))
    budget_object = BudgetTotals.objects.exclude(department_id__in=[13, 15, 16])
    budget_actual_id = [obj.account_id for obj in budget_object]
    budget_actuals = Glafs.objects.using('sql_server').filter(acctid__in=budget_actual_id, fscsdsg='A', fscscurn='ZMW',
                                                              curntype='F',
                                                              fscsyr=year.year)
    for budget_total, actual_total in zip(budget_object, budget_actuals):
        for netperd in range(1, 15):
            value_budget = getattr(budget_total, f'netperd{netperd}', Decimal('0'))
            value_actual = getattr(actual_total, f'netperd{netperd}', Decimal('0'))
            budget_totals[budget_total.account][f'netperd {netperd} (Budget)'] = value_budget
            budget_totals[budget_total.account][f'netperd {netperd} (Actual)'] = value_actual
            budget_totals[budget_total.account]['Account Name'] = budget_total.account.acctdesc
            for quarter in range(1, 5):
                start_netperd = (quarter - 1) * 3 + 1
                end_netperd = quarter * 3
                quarter_value_budget = sum(getattr(budget_total, f'netperd{netperd}', Decimal('0')) for netperd in
                                           range(start_netperd, end_netperd + 1))
                quarter_value_actual = sum(getattr(actual_total, f'netperd{netperd}', Decimal('0')) for netperd in
                                           range(start_netperd, end_netperd + 1))

                budget_totals[budget_total.account][f'Q{quarter} (Budget)'] = quarter_value_budget
                budget_totals[budget_total.account][f'Q{quarter} (Actual)'] = quarter_value_actual

    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active

    # Writing the header row with bold and blue background color
    for col, header in enumerate(excel_header, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = cell.font.copy(bold=True, color=colors.WHITE)
        cell.fill = cell.fill.copy(fill_type='solid', start_color='00008B')

    # Writing the data rows

    for row, (account, totals) in enumerate(budget_totals.items(), start=2):
        sheet.cell(row=row, column=1, value=account.pk)
        for col, netperd in enumerate(excel_header[1:], start=2):
            if 'Q' in netperd:
                sheet.cell(row=row, column=col, value=Decimal(totals[netperd]))
            elif 'netperd' in netperd:
                sheet.cell(row=row, column=col, value=Decimal(totals[netperd]))
            else:
                sheet.cell(row=row, column=col, value=totals[netperd])

    # Create a HttpResponse with an Excel file attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'inline; filename="budget-export-opex {datetime.today().date()}.xlsx"'

    # Save the workbook to the HttpResponse
    workbook.save(response)

    return response


def generate_excel_capex(request):
    # Define the header row for the Excel file
    excel_header = [
        'Account Number',
        'Account Name',
        'netperd 1 (Budget)',
        'netperd 1 (Actual)',
        'netperd 2 (Budget)',
        'netperd 2 (Actual)',
        'netperd 3 (Budget)',
        'netperd 3 (Actual)',
        'netperd 4 (Budget)',
        'netperd 4 (Actual)',
        'netperd 5 (Budget)',
        'netperd 5 (Actual)',
        'netperd 6 (Budget)',
        'netperd 6 (Actual)',
        'netperd 7 (Budget)',
        'netperd 7 (Actual)',
        'netperd 8 (Budget)',
        'netperd 8 (Actual)',
        'netperd 9 (Budget)',
        'netperd 9 (Actual)',
        'netperd 10 (Budget)',
        'netperd 10 (Actual)',
        'netperd 11 (Budget)',
        'netperd 11 (Actual)',
        'netperd 12 (Budget)',
        'netperd 12 (Actual)',
        'Q1 (Actual)',
        'Q1 (Budget)',
        'Q2 (Budget)',
        'Q2 (Actual)',
        'Q3 (Budget)',
        'Q3 (Actual)',
        'Q4 (Budget)',
        'Q4 (Actual)'
    ]
    year = FinancialYear.objects.get(is_active=True)

    budget_object = BudgetTotals.objects.filter(department_id=13)
    # Grouping the values according to quarters and halves
    budget_totals = defaultdict(lambda: defaultdict(Decimal))
    budget_actual_id = [obj.account_id for obj in budget_object]
    budget_actuals = Glafs.objects.using('sql_server').filter(acctid__in=budget_actual_id, fscsdsg='A', fscscurn='ZMW',
                                                              curntype='F',
                                                              fscsyr=year.year)
    for budget_total, actual_total in zip(budget_object, budget_actuals):
        for netperd in range(1, 13):
            value_budget = getattr(budget_total, f'netperd{netperd}', Decimal('0'))
            value_actual = getattr(budget_actuals, f'netperd{netperd}', Decimal('0'))
            budget_totals[budget_total.account][f'netperd {netperd} (Budget)'] = value_budget
            budget_totals[budget_total.account][f'netperd {netperd} (Actual)'] = value_actual
            budget_totals[budget_total.account]['Account Name'] = budget_total.account.acctdesc
            for quarter in range(1, 5):
                start_netperd = (quarter - 1) * 3 + 1
                end_netperd = quarter * 3
                quarter_value_budget = sum(getattr(budget_total, f'netperd{netperd}', Decimal('0')) for netperd in
                                           range(start_netperd, end_netperd + 1))
                quarter_value_actual = sum(getattr(budget_total, f'netperd{netperd}', Decimal('0')) for netperd in
                                           range(start_netperd, end_netperd + 1))

                budget_totals[budget_total.account][f'Q{quarter} (Budget)'] = quarter_value_budget
                budget_totals[budget_total.account][f'Q{quarter} (Actual)'] = quarter_value_budget

    # Create a new Excel workbook
    workbook = Workbook()
    sheet = workbook.active

    # Writing the header row with bold and blue background color
    for col, header in enumerate(excel_header, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = cell.font.copy(bold=True, color=colors.WHITE)
        cell.fill = cell.fill.copy(fill_type='solid', start_color='00008B')

    # Writing the data rows

    for row, (account, totals) in enumerate(budget_totals.items(), start=2):
        sheet.cell(row=row, column=1, value=account.pk)
        for col, netperd in enumerate(excel_header[1:], start=2):
            for col, netperd in enumerate(excel_header[1:], start=2):
                if 'Q' in netperd:
                    sheet.cell(row=row, column=col, value=Decimal(totals[netperd]))
                elif 'netperd' in netperd:
                    sheet.cell(row=row, column=col, value=Decimal(totals[netperd]))
                else:
                    sheet.cell(row=row, column=col, value=totals[netperd])

    # Create a HttpResponse with an Excel file attachment
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="budget-export-capex-{datetime.today().date()}.xlsx"'

    # Save the workbook to the HttpResponse
    workbook.save(response)

    return response


def department_budget_settings(request, dept_id):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)
        obj = BudgetStatus.objects.filter(department_id=dept_id, year=year.year).all()

        if request.GET.get('complete') == '1':
            if request.user.role != '002':
                message = f"{request.GET.get('budget_id')} from {request.user.department.name} department has been marked as complete"
                # Send the notification message
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    'public_room',
                    {
                        "type": "send_notification",
                        "message": message
                    }
                )
            BudgetStatus.objects.filter(department=request.user.department,
                                        budget_set=request.GET.get('budget_id'), year=year.year).update(
                is_complete=True,
                comment=request.GET.get(
                    'comment'))
            mail_list = Users.objects.filter(email='seriterk@gmail.com').values('email','first_name')
            for user in mail_list:
                receiver_email = user['email']
                receiver_name = user['first_name']

                # Prepare the context for rendering the email template
                context = {
                    "receiver": receiver_name,  # Use 'receiver' as per your template variable
                    "message": message,
                }

                # Define the template name for the HTML email
                template_name = "templates/status-change-mail.html"

                # Render the HTML content from the template
                convert_to_html_content = render_to_string(template_name, context)

                # Strip HTML tags for plain text version of the email
                plain_message = strip_tags(convert_to_html_content)

                # Send the email
                yo_send_it = send_mail(
                    subject="Budget Module Notification",
                    message=plain_message,
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[receiver_email],  # Send to one recipient at a time
                    html_message=convert_to_html_content,
                    fail_silently=True  # Optional: Set to False to raise errors during sending
                )
            return JsonResponse({'result': 'success'})
        elif request.GET.get('incomplete') == '1':
            if request.user.role != '002':
                message = f"Changes to be made to {request.GET.get('budget_id')} from {request.user.department.name} department require your approval"

                # Send the notification message
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    'public_room',
                    {
                        "type": "send_notification",
                        "message": message
                    }
                )
            budget = request.GET.get('budget_id')
            BudgetStatus.objects.filter(department=request.user.department, budget_set=budget, year=year.year).update(
                is_complete=False,
                comment=request.GET.get(
                    'comment'))
            return JsonResponse({'result': 'success'})
        return render(request, 'budget-settings.html', {'budgetStatus': obj})
    else:
        return redirect('budgets:login')


def budget_settings(request):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)

        obj = BudgetStatus.objects.filter(year=year.year).all()
        if request.GET.get('complete') == '1':
            if request.user.role != '002':
                message = f"{request.GET.get('budget_id')} from {request.user.department.name} department has been marked as complete"
                # Send the notification message
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    'public_room',
                    {
                        "type": "send_notification",
                        "message": message
                    }
                )
            BudgetStatus.objects.filter(department=request.user.department,
                                        budget_set=request.GET.get('budget_id'), year=year.year).update(
                is_complete=True,

                comment=request.GET.get(
                    'comment'))
            return JsonResponse({'result': 'success'})
        elif request.GET.get('incomplete') == '1':
            if request.user.role != '002':
                message = f"Changes to be made to {request.GET.get('budget_id')} from {request.user.department.name} department require your approval"

                # Send the notification message
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    'public_room',
                    {
                        "type": "send_notification",
                        "message": message
                    }
                )
            budget = request.GET.get('budget_id')
            BudgetStatus.objects.filter(department=request.user.department, budget_set=budget, year=year.year).update(
                is_complete=False,
                comment=request.GET.get(
                    'comment'))
            return JsonResponse({'result': 'success'})
        toggle_init = ChangeLog.objects.get(year=year.year)
        toggle = False
        if toggle_init.flag == True:
            toggle=True
        return render(request, 'budget-settings.html', {'budgetStatus': obj, 'toggle':toggle})
    else:
        return redirect('budgets:login')


def budget_assumptions(request):
    if request.user.is_authenticated:
        obj = BudgetAssumptions.objects.filter(department_id=request.user.department.id)
        currency_obj = Currency.objects.all()
        if request.method == 'POST':
            data = json.loads(request.body)

            if data['target'] == "assumptions":

                assumption_obj = BudgetAssumptions.objects.filter(factor=data['factor'],
                                                                  department=request.user.department).first()

                if not assumption_obj:
                    BudgetAssumptions.objects.create(rate=data['rate'], factor=data['factor'],
                                                     department=request.user.department, updated_by=request.user)
                else:

                    assumption_obj.rate = data['rate']
                    assumption_obj.save()

                obj = BudgetAssumptions.objects.filter(
                    department_id=request.user.department.id)  # Refresh assumptions
            elif data['target'] == 'currency':
                currency = Currency.objects.filter(currency=data['currency']).first()
                if not currency:
                    Currency.objects.create(currency=data['currency'], rate=data['rate'], updated_by=request.user)
                else:
                    currency.rate = data['rate']
                    currency.save()
                currency_obj = Currency.objects.all()
        paginator = Paginator(obj, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context = {'assumptions': page_obj, 'currency': currency_obj, 'department': request.user.department.name}
        return render(request, 'assumptions.html', context)
    else:
        return redirect('budgets:login')


def accounts_search(request):
    year = FinancialYear.objects.get(is_active=True)
    obj = BudgetVariations.objects.get(is_active=True,year=year.year)
    field_mapping = {
        'account_id': 'account__acctid__icontains',
        'account_name': 'account__acctdesc__icontains',
    }

    filter = request.GET.get('filter')
    value = request.GET.get('value')
    dept = request.GET.get('department')
    active = BudgetVariations.objects.filter(is_active=True).values('budget_set')
    if dept:
        account_info = BudgetTotals.objects.filter(budget_set=obj.budget_set, **{field_mapping[filter]: value},
                                                   department_id=dept, year=year.year).values('id', 'total',
                                                                                              'account_id',
                                                                                              'account__acctdesc',
                                                                                              'year',
                                                                                              'currency__currency',
                                                                                              'netperd1',
                                                                                              'netperd2', 'netperd3',
                                                                                              'netperd4',
                                                                                              'netperd5', 'netperd6',
                                                                                              'netperd7',
                                                                                              'netperd8', 'netperd9',
                                                                                              'netperd10',
                                                                                              'netperd11',
                                                                                              'netperd12').all()
        print(account_info)
        return JsonResponse({'data': list(account_info)}, status=200)
    else:
        if filter in field_mapping and value:
            if request.user.role != '002':
                account_info = BudgetTotals.objects.filter(budget_set=obj.budget_set,
                                                           **{field_mapping[filter]: value},
                                                           department=request.user.department, year=year.year).values(
                    'account_id', 'account__acctdesc', 'id').all()
                return JsonResponse({'data': list(account_info)}, status=200)
            else:
                account_info = BudgetTotals.objects.filter(budget_set=obj.budget_set,
                                                           **{field_mapping[filter]: value}, year=year.year).values(
                    'account_id', 'account__acctdesc', 'id').all()
                return JsonResponse({'data': list(account_info)}, status=200)


def get_budget_set(request, object_id, budget_set):
    year = FinancialYear.objects.get(is_active=True)
    single_object = BudgetTotals.objects.all()
    obj = get_object_or_404(BudgetTotals, id=object_id, budget_set=budget_set, year=year.year)

    next_obj = BudgetTotals.objects.filter(id__gt=obj.id).order_by('id').first()
    prev_obj = BudgetTotals.objects.filter(id__lt=obj.id).order_by('-id').first()

    first_obj = BudgetTotals.objects.first()
    last_obj = BudgetTotals.objects.last()
    form = BudgetEditForm(instance=obj)
    context = {
        'form': form,
        'obj': single_object,
        'object': obj,
        'next_obj': next_obj,
        'prev_obj': prev_obj,
        'first_obj': first_obj,
        'last_obj': last_obj
    }

    return render(request, 'forms.html', context)


def user_logout(request):
    logout(request)
    return redirect('budgets:login')


def check_otp(entered_otp, saved_otp):
    if entered_otp == saved_otp:
        return True
    else:
        return False


def verify_otp(request):
    if request.method == 'POST':
        otp = request.POST.get('otp')
        user_id = request.user.id

        user = get_object_or_404(Users, id=user_id)

        if check_password(otp, user.login_otp) or user.is_valid_otp():
            RefreshToken.for_user(user)
            user.login_otp_used = True
            user.save(update_fields=["login_otp_used"])
            return redirect('budgets:home', budget_set=1)
        else:
            messages.error(request, message="Authentication Failed Try Again")
            return redirect("budgets:login")
    return render(request, 'enter-otp.html')


def create_user(request):
    form = UserCreate(request.POST)

    if form.is_valid():
        form.save()
        return redirect('budgets:login')
    return render(request, 'user-create.html', {'form': UserCreate(request.GET)})


def toggle_status_incomplete(request, id):
    budget_obj = BudgetStatus.objects.get(id=id)
    budget_obj.is_complete = False
    budget_obj.save()
    return redirect('budgets:settings')


def toggle_status_complete(request, id):
    budget_obj = BudgetStatus.objects.get(id=id)
    budget_obj.is_complete = True
    budget_obj.save()
    return redirect('budgets:settings')


def toggle_status_false(request, id):
    year = FinancialYear.objects.get(is_active=True)
    BudgetStatus.objects.filter(budget_set=id, year=year.year).update(is_active=False)
    BudgetVariations.objects.filter(budget_set=id, year=year.year).update(is_active=False)

    return redirect('budgets:settings')


def toggle_status_true(request, id):
    try:
        year = FinancialYear.objects.get(is_active=True)
        BudgetStatus.objects.filter(budget_set=id, year=year.year).update(is_active=True)
        BudgetStatus.objects.exclude(budget_set=id, year=year.year).update(is_active=False)
        BudgetVariations.objects.filter(budget_set=id, year=year.year).update(is_active=True)
        BudgetVariations.objects.exclude(budget_set=id, year=year.year).update(is_active=False)
        return redirect('budgets:settings')
    except ValidationError as e:
        messages.error(request, e.message)
        return redirect('budgets:settings')


def clear_budget_line(request, object_id):
    line = get_object_or_404(BudgetLines, id=object_id)
    active = get_object_or_404(BudgetVariations, is_active=True)
    year = FinancialYear.objects.get(is_active=True)
    dept = get_object_or_404(BudgetTotals, account_id=line.account_id, budget_set=active.budget_set, year=year.year)
    if dept:
        with transaction.atomic():
            BudgetTotals.objects.filter(account_id=line.account.acctid, budget_set=active.budget_set,
                                        year=year.year).update(
                total=F('total') - line.total,
                netperd1=F('netperd1') - line.netperd1,
                netperd2=F('netperd2') - line.netperd2,
                netperd3=F('netperd3') - line.netperd3,
                netperd4=F('netperd4') - line.netperd4,
                netperd5=F('netperd5') - line.netperd5,
                netperd6=F('netperd6') - line.netperd6,
                netperd7=F('netperd7') - line.netperd7,
                netperd8=F('netperd8') - line.netperd8,
                netperd9=F('netperd9') - line.netperd9,
                netperd10=F('netperd10') - line.netperd10,
                netperd11=F('netperd11') - line.netperd11,
                netperd12=F('netperd12') - line.netperd12,
                last_updated=datetime.now(),
                last_updated_by_id=request.user.id
            )
            line.delete()

    return redirect('budgets:update', dept.id)




def clear_budget_numbers(request, object_id):
    active = get_object_or_404(BudgetVariations, active=True)
    line = get_object_or_404(BudgetTotals, id=object_id)
    year = FinancialYear.objects.get(is_active=True)
    if line:
        BudgetLines.objects.filter(account_id=line.account.acctid, budget_set=active.budget_set,
                                   year=year.year).delete()
        with transaction.atomic():
            BudgetTotals.objects.filter(account_id=line.account.acctid, budget_set=active.budget_set,
                                        year=year.year).update(
                total=F('total') - line.total,
                netperd1=F('netperd1') - line.netperd1,
                netperd2=F('netperd2') - line.netperd2,
                netperd3=F('netperd3') - line.netperd3,
                netperd4=F('netperd4') - line.netperd4,
                netperd5=F('netperd5') - line.netperd5,
                netperd6=F('netperd6') - line.netperd6,
                netperd7=F('netperd7') - line.netperd7,
                netperd8=F('netperd8') - line.netperd8,
                netperd9=F('netperd9') - line.netperd9,
                netperd10=F('netperd10') - line.netperd10,
                netperd11=F('netperd11') - line.netperd11,
                netperd12=F('netperd12') - line.netperd12,
                last_updated=datetime.now(),
                last_updated_by_id=request.user.id
            )

    return redirect('budgets:dashboard-home-department')


def changelog(request):
    if request.user.is_authenticated:
        obj = BudgetLinesLog.objects.all()

        year = FinancialYear.objects.get(is_active=True)
        obj = BudgetVariations.objects.get(is_active=True, year=year.year)
        # Combine all department conditions using Q objects
        department_conditions = Q()
        for department_id in range(1, 15):  # Assuming departments IDs range from 1 to 12
            department_conditions |= Q(department_id=department_id)

        # Fetch all relevant budget totals in a single query
        budget_totals = BudgetLinesLog.objects.filter(year=year.year).order_by('-account_id')

        # Create separate variables for each department
        ceo = budget_totals.filter(department_id=1)
        paginator = Paginator(ceo, 10)
        page_number = request.GET.get('page')
        ceo_obj = paginator.get_page(page_number)
        internal_audit = budget_totals.filter(department_id=2)
        paginator = Paginator(internal_audit, 10)
        page_number = request.GET.get('page')
        internal_audit_obj = paginator.get_page(page_number)
        supply_chain = budget_totals.filter(department_id=3)
        paginator = Paginator(supply_chain, 10)
        page_number = request.GET.get('page')
        supply_chain_obj = paginator.get_page(page_number)
        bds = budget_totals.filter(department_id=4)
        paginator = Paginator(bds, 10)
        page_number = request.GET.get('page')
        bds_obj = paginator.get_page(page_number)
        public_relations = budget_totals.filter(department_id=5)
        paginator = Paginator(public_relations, 10)
        page_number = request.GET.get('page')
        public_relations_obj = paginator.get_page(page_number)
        technical = budget_totals.filter(department_id=6)
        paginator = Paginator(technical, 10)
        page_number = request.GET.get('page')
        technical_obj = paginator.get_page(page_number)
        information_systems = budget_totals.filter(department_id=7)
        paginator = Paginator(information_systems, 10)
        page_number = request.GET.get('page')
        information_systems_obj = paginator.get_page(page_number)
        legal_risk = budget_totals.filter(department_id=8)
        paginator = Paginator(legal_risk, 10)
        page_number = request.GET.get('page')
        legal_risk_obj = paginator.get_page(page_number)
        human_capital = budget_totals.filter(department_id=9)
        paginator = Paginator(human_capital, 10)
        page_number = request.GET.get('page')
        human_capital_obj = paginator.get_page(page_number)
        sales_marketing = budget_totals.filter(department_id=10)
        paginator = Paginator(sales_marketing, 10)
        page_number = request.GET.get('page')
        sales_marketing_obj = paginator.get_page(page_number)
        admin = budget_totals.filter(department_id=11)
        paginator = Paginator(admin, 10)
        page_number = request.GET.get('page')
        admin_page_obj = paginator.get_page(page_number)
        finance = budget_totals.filter(department_id=12)
        paginator = Paginator(finance, 10)
        page_number = request.GET.get('page')
        finance_page_obj = paginator.get_page(page_number)
        staff_remuneration = budget_totals.filter(department_id=14)
        paginator = Paginator(staff_remuneration, 10)
        page_number = request.GET.get('page')
        staff_n_page_obj = paginator.get_page(page_number)
        assets = budget_totals.filter(department_id=13)
        paginator = Paginator(assets, 10)
        page_number = request.GET.get('page')
        assets = paginator.get_page(page_number)
        # Calculate total sum
        # Populate context variable
        context = {
            'staff_remuneration': staff_n_page_obj,
            'ceo': ceo_obj,
            'internal_audit': internal_audit_obj,
            'supply_chain': supply_chain_obj,
            'bds': bds_obj,
            'public_relations': public_relations_obj,
            'technical': technical_obj,
            'information_systems': information_systems_obj,
            'legal_risk': legal_risk_obj,
            'human_capital': human_capital_obj,
            'sales_marketing': sales_marketing_obj,
            'admin': admin_page_obj,
            'finance': finance_page_obj,
            'assets': assets,
            'budget_set': obj.budget_set,
        }

        return render(request, 'changelog.html', context)
    else:
        return redirect('budgets:login')
def variationchangelog(request):
    if request.user.is_authenticated:


        year = FinancialYear.objects.get(is_active=True)
        obj = BudgetVariations.objects.get(is_active=True, year=year.year)
        # Combine all department conditions using Q objects
        department_conditions = Q()
        for department_id in range(1, 15):  # Assuming departments IDs range from 1 to 12
            department_conditions |= Q(department_id=department_id)

        # Fetch all relevant budget totals in a single query
        budget_totals = BudgetLinesLogVariations.objects.filter(year=year.year).order_by('-account_id')

        # Create separate variables for each department
        ceo = budget_totals.filter(department_id=1)
        paginator = Paginator(ceo, 10)
        page_number = request.GET.get('page')
        ceo_obj = paginator.get_page(page_number)
        internal_audit = budget_totals.filter(department_id=2)
        paginator = Paginator(internal_audit, 10)
        page_number = request.GET.get('page')
        internal_audit_obj = paginator.get_page(page_number)
        supply_chain = budget_totals.filter(department_id=3)
        paginator = Paginator(supply_chain, 10)
        page_number = request.GET.get('page')
        supply_chain_obj = paginator.get_page(page_number)
        bds = budget_totals.filter(department_id=4)
        paginator = Paginator(bds, 10)
        page_number = request.GET.get('page')
        bds_obj = paginator.get_page(page_number)
        public_relations = budget_totals.filter(department_id=5)
        paginator = Paginator(public_relations, 10)
        page_number = request.GET.get('page')
        public_relations_obj = paginator.get_page(page_number)
        technical = budget_totals.filter(department_id=6)
        paginator = Paginator(technical, 10)
        page_number = request.GET.get('page')
        technical_obj = paginator.get_page(page_number)
        information_systems = budget_totals.filter(department_id=7)
        paginator = Paginator(information_systems, 10)
        page_number = request.GET.get('page')
        information_systems_obj = paginator.get_page(page_number)
        legal_risk = budget_totals.filter(department_id=8)
        paginator = Paginator(legal_risk, 10)
        page_number = request.GET.get('page')
        legal_risk_obj = paginator.get_page(page_number)
        human_capital = budget_totals.filter(department_id=9)
        paginator = Paginator(human_capital, 10)
        page_number = request.GET.get('page')
        human_capital_obj = paginator.get_page(page_number)
        sales_marketing = budget_totals.filter(department_id=10)
        paginator = Paginator(sales_marketing, 10)
        page_number = request.GET.get('page')
        sales_marketing_obj = paginator.get_page(page_number)
        admin = budget_totals.filter(department_id=11)
        paginator = Paginator(admin, 10)
        page_number = request.GET.get('page')
        admin_page_obj = paginator.get_page(page_number)
        finance = budget_totals.filter(department_id=12)
        paginator = Paginator(finance, 10)
        page_number = request.GET.get('page')
        finance_page_obj = paginator.get_page(page_number)
        staff_remuneration = budget_totals.filter(department_id=14)
        paginator = Paginator(staff_remuneration, 10)
        page_number = request.GET.get('page')
        staff_n_page_obj = paginator.get_page(page_number)
        assets = budget_totals.filter(department_id=13)
        paginator = Paginator(assets, 10)
        page_number = request.GET.get('page')
        assets = paginator.get_page(page_number)
        # Calculate total sum
        # Populate context variable
        context = {
            'staff_remuneration': staff_n_page_obj,
            'ceo': ceo_obj,
            'internal_audit': internal_audit_obj,
            'supply_chain': supply_chain_obj,
            'bds': bds_obj,
            'public_relations': public_relations_obj,
            'technical': technical_obj,
            'information_systems': information_systems_obj,
            'legal_risk': legal_risk_obj,
            'human_capital': human_capital_obj,
            'sales_marketing': sales_marketing_obj,
            'admin': admin_page_obj,
            'finance': finance_page_obj,
            'assets': assets,
            'budget_set': obj.budget_set,
        }

        return render(request, 'variationlog.html', context)
    else:
        return redirect('budgets:login')


def update(request, object_id):
    if request.user.is_authenticated:
        year = FinancialYear.objects.get(is_active=True)
        obj_t = get_object_or_404(BudgetTotals, id=object_id, year=year.year)
        lines = BudgetLines.objects.filter(account__acctid=obj_t.account.acctid, year=year.year)
    
        active = get_object_or_404(BudgetVariations, is_active=True,year=year.year)
        if active:
            if request.user.role != '002':
                single_object = BudgetTotals.objects.filter(budget_set=active.budget_set,
                                                            department=request.user.department, year=year.year)
            else:
                single_object = BudgetTotals.objects.filter(budget_set=active.budget_set, year=year.year)

            first_dept_obj = BudgetTotals.objects.filter(budget_set=active.budget_set,
                                                         department_id=obj_t.department.id,
                                                         id=object_id, year=year.year).first()
            user = get_object_or_404(Users, id=request.user.id)
            obj = get_object_or_404(BudgetTotals, id=first_dept_obj.id, year=year.year)
            if request.user.role != '002':
                assumptions = BudgetAssumptions.objects.filter(department=request.user.department)
            else:
                assumptions = BudgetAssumptions.objects.all()
            next_obj = BudgetTotals.objects.filter(id__gt=obj.id, department_id=obj_t.department.id,
                                                   budget_set=active.budget_set, year=year.year).order_by('id').first()
            prev_obj = BudgetTotals.objects.filter(id__lt=obj.id, department_id=obj_t.department.id,
                                                   budget_set=active.budget_set, year=year.year).order_by('-id').first()

            first_obj = BudgetTotals.objects.filter(department_id=obj_t.department.id,
                                                    budget_set=active.budget_set, year=year.year).first()
            last_obj = BudgetTotals.objects.filter(department_id=obj_t.department.id,
                                                   budget_set=active.budget_set, year=year.year).last()

            form = BudgetEditForm()

            if request.method == 'POST':

                data = json.loads(request.body)
                currency_obj = get_object_or_404(Currency, id=data['currency'])
                with transaction.atomic():
                    BudgetTotals.objects.filter(id=object_id, year=year.year).update(last_updated=datetime.now(),
                                                                                     last_updated_by_id=request.user.id,

                                                                                     total=F('total') + data['total'],
                                                                                     netperd1=F('netperd1') + data[
                                                                                         'netperd1'],
                                                                                     netperd2=F('netperd2') + data[
                                                                                         'netperd2'],
                                                                                     netperd3=F('netperd3') + data[
                                                                                         'netperd3'],
                                                                                     netperd4=F('netperd4') + data[
                                                                                         'netperd4'],
                                                                                     netperd5=F('netperd5') + data[
                                                                                         'netperd5'],
                                                                                     netperd6=F('netperd6') + data[
                                                                                         'netperd6'],
                                                                                     netperd7=F('netperd7') + data[
                                                                                         'netperd7'],
                                                                                     netperd8=F('netperd8') + data[
                                                                                         'netperd8'],
                                                                                     netperd9=F('netperd9') + data[
                                                                                         'netperd9'],
                                                                                     netperd10=F('netperd10') + data[
                                                                                         'netperd10'],
                                                                                     netperd11=F('netperd11') + data[
                                                                                         'netperd11'],
                                                                                     netperd12=F('netperd12') + data[
                                                                                         'netperd12']
                                                                                     )
                if data['entryType'] == 'function':
                    assumption_obj = get_object_or_404(BudgetAssumptions, factor=data['assumption'])
                    BudgetLines.objects.create(
                        year=year.year,
                        account=obj.account,
                        exchange_rate=data['exchange_rate'],
                        item_description=data['item_description'],
                        department=obj.department,
                        total=Decimal(data['total']),
                        rate=Decimal(data['rate']),
                        usage=Decimal(data['usage']),
                        factor=Decimal(data['factor']),
                        staff=Decimal(data['staff']),
                        netperd1=Decimal(data['netperd1']),
                        netperd2=Decimal(data['netperd2']),
                        netperd3=Decimal(data['netperd3']),
                        netperd4=Decimal(data['netperd4']),
                        netperd5=Decimal(data['netperd5']),
                        netperd6=Decimal(data['netperd6']),
                        netperd7=Decimal(data['netperd7']),
                        netperd8=Decimal(data['netperd8']),
                        netperd9=Decimal(data['netperd9']),
                        netperd10=Decimal(data['netperd10']),
                        netperd11=Decimal(data['netperd11']),
                        netperd12=Decimal(data['netperd12']),
                        currency=currency_obj,
                        assumption=assumption_obj,
                        budget_set=active.budget_set,
                        last_updated=datetime.now(),
                        last_updated_by_id=request.user.id)
                    return JsonResponse({'status': 'success'})
                elif data['entryType'] == 'manual':

                    BudgetLines.objects.create(
                        account=obj.account,
                        year=year.year,
                        exchange_rate=data['exchange_rate'],
                        department=obj.department,
                        total=Decimal(data['total']),
                        item_description=data['item_description'],
                        rate=Decimal(0),
                        usage=Decimal(0),
                        factor=Decimal(0),
                        staff=Decimal(0),
                        netperd1=Decimal(data['netperd1']),
                        netperd2=Decimal(data['netperd2']),
                        netperd3=Decimal(data['netperd3']),
                        netperd4=Decimal(data['netperd4']),
                        netperd5=Decimal(data['netperd5']),
                        netperd6=Decimal(data['netperd6']),
                        netperd7=Decimal(data['netperd7']),
                        netperd8=Decimal(data['netperd8']),
                        netperd9=Decimal(data['netperd9']),
                        netperd10=Decimal(data['netperd10']),
                        netperd11=Decimal(data['netperd11']),
                        netperd12=Decimal(data['netperd12']),
                        currency=currency_obj,
                        budget_set=active.budget_set,
                        last_updated=datetime.now(),
                        last_updated_by_id=request.user.id)

                    return JsonResponse({'status': 'success'})
            context = {
                'currency': Currency.objects.all(),
                'name': obj.account.acctdesc,
                'form': form,
                'obj': single_object,
                'object': obj,
                'next_obj': next_obj,
                'prev_obj': prev_obj,
                'first_obj': first_obj,
                'last_obj': last_obj,
                'assumptions': assumptions,
                'activebs': active.budget_set,
                'department': obj_t.department.name,
                'lines': lines,
                'budget_set': active.budget_set,

            }

            return render(request, 'forms.html', context)
    else:
        return redirect('budgets:login')

def is_numeric(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

def vary_lines(request):
    year = FinancialYear.objects.get(is_active=True)
    active = get_object_or_404(BudgetVariations, is_active=True, year=year.year)

    form = json.loads(request.body)
    if is_numeric(form.get('to_line')):
        amount = Decimal(form.get('from_amount'))
        period = form.get('to_period')
        fromPeriod = form.get('from_period')

        fromLine = get_object_or_404(BudgetLines, id=form.get('from_line'), year=year.year)
        fromTotal = get_object_or_404(BudgetTotals, account_id=fromLine.account_id, budget_set=active.budget_set,
                                      year=year.year)

        # Check if we're updating the same line
        same_line = str(form.get('from_line')) == str(form.get('to_line'))

        if not same_line:
            toLine = get_object_or_404(BudgetLines, id=form.get('to_line'), year=year.year)
            toTotal = get_object_or_404(BudgetTotals, account_id=toLine.account_id, budget_set=active.budget_set,
                                        year=year.year)
        else:
            toLine = fromLine
            toTotal = fromTotal

        current_timestamp = datetime.now()
        fromLine.timestamp = current_timestamp
        fromTotal.timestamp = current_timestamp
        toLine.timestamp = current_timestamp
        toTotal.timestamp = current_timestamp

        # For same line updates, we need to handle the calculations differently
        if same_line and period == fromPeriod:
            # If moving within the same period of the same line, no action needed
            return JsonResponse({'message': 'success'}, status=200)

        # Calculate all changes before saving
        changes = {}

        # Calculate changes for the source period
        netperd_field_from = f'{fromPeriod}'
        if hasattr(fromLine, netperd_field_from):
            current_value = getattr(fromLine, netperd_field_from, 0)
            changes['from_value'] = current_value - amount
            # Only adjust the total if it's not the same line
            if not same_line:
                changes['from_total'] = fromLine.total - amount
            else:
                changes['from_total'] = fromLine.total  # Keep the total unchanged for same line

        # Calculate changes for the destination period
        netperd_field_to = f'{period}'
        if hasattr(toLine, netperd_field_to):
            current_value = getattr(toLine, netperd_field_to, 0)
            changes['to_value'] = current_value + amount
            # Only adjust the total if it's not the same line
            if not same_line:
                changes['to_total'] = toLine.total + amount
            else:
                changes['to_total'] = toLine.total  # Keep the total unchanged for same line

        # Apply all changes
        if 'from_value' in changes:
            setattr(fromLine, netperd_field_from, changes['from_value'])
            fromLine.total = changes['from_total']

            if hasattr(fromTotal, netperd_field_from):
                current_value = getattr(fromTotal, netperd_field_from, 0)
                setattr(fromTotal, netperd_field_from, current_value - amount)
            if hasattr(fromTotal, 'total') and not same_line:
                current_value = getattr(fromTotal, 'total', 0)
                setattr(fromTotal, 'total', current_value - amount)

        if 'to_value' in changes:
            setattr(toLine, netperd_field_to, changes['to_value'])
            toLine.total = changes['to_total']

            if hasattr(toTotal, netperd_field_to):
                current_value = getattr(toTotal, netperd_field_to, 0)
                setattr(toTotal, netperd_field_to, current_value + amount)
            if hasattr(toTotal, 'total') and not same_line:
                current_value = getattr(toTotal, 'total', 0)
                setattr(toTotal, 'total', current_value + amount)

        # Save all changes
        fromLine.save()
        fromTotal.save()
        if not same_line:
            toLine.save()
            toTotal.save()

        # Create log entries
        BudgetLinesLogVariations.objects.create(
            timestamp=current_timestamp,
            user=request.user,
            action="removed",
            item_description=fromLine.item_description,
            account=fromLine.account,
            department=fromLine.department,
            period=form.get('from_period'),
            amount=-amount,
            year=year.year,
            comment=form.get('comment')
        )

        BudgetLinesLogVariations.objects.create(
            timestamp=current_timestamp,
            user=request.user,
            action="added",
            item_description=toLine.item_description,
            account=toLine.account,
            department=toLine.department,
            period=form.get('to_period'),
            amount=amount,
            year=year.year,
            comment=form.get('comment')
        )
    else:
        amount = form.get('from_amount')
        account = form.get('to_account')
        period = form.get('to_period')
        fromPeriod = form.get('from_period')

        fromLine = get_object_or_404(BudgetLines, id=form.get('from_line'), year=year.year)
        fromTotal = get_object_or_404(BudgetTotals, account_id=fromLine.account_id, budget_set=active.budget_set,
                                      year=year.year)

        fromLine.timestamp = datetime.now()
        fromTotal.timestamp = datetime.now()


        # Updates values of affected periods of from account
        netperd_field_from = f'{fromPeriod}'  # Create the field name dynamically, e.g., 'netperd1', 'netperd2', etc.
        if hasattr(fromLine, netperd_field_from):
            current_value = getattr(fromLine, netperd_field_from, 0)  # Get the current value, default to 0 if None
            new_value = current_value - Decimal(amount)  # Subtract the amount from the current value

            setattr(fromLine, netperd_field_from, new_value)
            fromLine.total = fromLine.total - Decimal(amount)

        if hasattr(fromTotal, netperd_field_from):
            current_value = getattr(fromTotal, netperd_field_from, 0)
            new_value = current_value - Decimal(amount)
            setattr(fromTotal, netperd_field_from, new_value)
        if hasattr(fromTotal, 'total'):
            current_value = getattr(fromTotal, 'total', 0)
            new_value = current_value - Decimal(amount)
            setattr(fromTotal, 'total', new_value)
        fromLine.save()
        fromTotal.save()

        BudgetLinesLogVariations.objects.create(
            timestamp=datetime.now(),
            user=request.user,
            action="removed",
            item_description=fromLine.item_description,
            account=fromLine.account,
            department=fromLine.department,
            period=form.get('from_period'),
            amount=-Decimal(form.get('from_amount')),
            year=year.year,
            comment = form.get('comment')
        )

        acc = Accounts.objects.get(acctid=account)
        netperd_fields = {f'netperd{i}': Decimal(0) for i in range(1, 13)}
        update_fields = {f'netperd{i}': Decimal(0) for i in range(1, 13)}

        # Assign the amount to the specific period
        netperd_fields[form.get('to_period')] = amount
        update_fields[form.get('to_period')] = F(period) + amount
        BudgetTotals.objects.filter(account_id = account, year=year.year).update(
            last_updated=datetime.now(),
            last_updated_by_id=request.user.id,
            total=F('total') + amount,  # Update total amount
            **update_fields  # Dynamically pass all period updates
        )
        BudgetLines.objects.create(
            account_id=account,
            year=year.year,
            exchange_rate=1.0,
            department_id=acc.department_id,
            total=amount,
            item_description=form.get('to_line'),
            rate=Decimal(0),
            usage=Decimal(0),
            factor=Decimal(0),
            staff=Decimal(0),
            **netperd_fields,
            currency_id=1,
            budget_set=active.budget_set,
            last_updated=datetime.now(),
            last_updated_by_id=request.user.id)

        BudgetLinesLogVariations.objects.create(
                timestamp=datetime.now(),
                user=request.user,
                action="added",
                item_description=form.get('to_line'),
                account=acc,
                department_id=acc.department_id,
                period=form.get('to_period'),
                amount=Decimal(form.get('from_amount')),
                year=year.year,
                comment = form.get('comment')
        )


    return JsonResponse({'message': 'success'}, status=200)



def log_budget_line_action(budget_line, action, user):
    BudgetLinesLog.objects.create(
        user=user,
        action=action,
        amount=budget_line.total,  # You can log a specific amount or total
        item_description=budget_line.item_description,
        account=budget_line.account,
        total=budget_line.total,
        netperd1=budget_line.netperd1,
        netperd2=budget_line.netperd2,
        netperd3=budget_line.netperd3,
        netperd4=budget_line.netperd4,
        netperd5=budget_line.netperd5,
        netperd6=budget_line.netperd6,
        netperd7=budget_line.netperd7,
        netperd8=budget_line.netperd8,
        netperd9=budget_line.netperd9,
        netperd10=budget_line.netperd10,
        netperd11=budget_line.netperd11,
        netperd12=budget_line.netperd12,
        department = budget_line.department,
        year=budget_line.year
    )

def edit_line(request, object_id):
    if request.user.is_authenticated:

        year = FinancialYear.objects.get(is_active=True)
        active = get_object_or_404(BudgetVariations, is_active=True,year =year.year)
        line = get_object_or_404(BudgetLines, id=object_id, year=year.year)
        obj_t = get_object_or_404(BudgetTotals, account__acctid=line.account.acctid, budget_set=active.budget_set,
                                  year=year.year)
        lines = BudgetLines.objects.filter(account__acctid=obj_t.account.acctid, year=year.year)

        if active:

            if request.user.role != '002':
                assumptions = BudgetAssumptions.objects.filter(department=request.user.department)
            else:
                assumptions = BudgetAssumptions.objects.all()

            form = BudgetEditForm(instance=line)

            if request.method == 'POST':
                data = json.loads(request.body)
                currency_obj = get_object_or_404(Currency, id=data['currency'])
                if data['entryType'] == 'function':
                    assumption_obj = get_object_or_404(BudgetAssumptions, factor=data['assumption'])
                    with transaction.atomic():  # Ensures that all operations are atomic
                        budget_line = BudgetLines.objects.get(
                            pk=object_id,
                            year=year.year, )  # Assuming you have the ID of the BudgetLines instance to update
                        budget_line.account = obj_t.account
                        budget_line.exchange_rate = data['exchange_rate']
                        budget_line.item_description = data['item_description']
                        budget_line.department = obj_t.department
                        budget_line.total = Decimal(data['total'])
                        budget_line.rate = Decimal(data['rate'])
                        budget_line.usage = Decimal(data['usage'])
                        budget_line.factor = Decimal(data['factor'])
                        budget_line.staff = Decimal(data['staff'])
                        budget_line.netperd1 = Decimal(data['netperd1'])
                        budget_line.netperd2 = Decimal(data['netperd2'])
                        budget_line.netperd3 = Decimal(data['netperd3'])
                        budget_line.netperd4 = Decimal(data['netperd4'])
                        budget_line.netperd5 = Decimal(data['netperd5'])
                        budget_line.netperd6 = Decimal(data['netperd6'])
                        budget_line.netperd7 = Decimal(data['netperd7'])
                        budget_line.netperd8 = Decimal(data['netperd8'])
                        budget_line.netperd9 = Decimal(data['netperd9'])
                        budget_line.netperd10 = Decimal(data['netperd10'])
                        budget_line.netperd11 = Decimal(data['netperd11'])
                        budget_line.netperd12 = Decimal(data['netperd12'])
                        budget_line.currency = currency_obj
                        budget_line.assumption = assumption_obj
                        budget_line.budget_set = active.budget_set
                        budget_line.last_updated = datetime.now()
                        budget_line.last_updated_by_id = request.user.id
                        budget_line.save()  # Call save() to trigger pre-save signal


                    return redirect('budgets:update', obj_t.id)

                elif data['entryType'] == 'manual':
                    with transaction.atomic():  # Ensures that all operations are atomic
                        budget_line = BudgetLines.objects.get(
                            pk=object_id,
                            year=year.year)  # Assuming you have the ID of the BudgetLines instance to update
                        budget_line.account = obj_t.account
                        budget_line.exchange_rate = data['exchange_rate']
                        budget_line.department = obj_t.department
                        budget_line.total = Decimal(data['total'])
                        budget_line.item_description = data['item_description']
                        budget_line.rate = Decimal(0)
                        budget_line.usage = Decimal(0)
                        budget_line.factor = Decimal(0)
                        budget_line.staff = Decimal(0)

                        budget_line.netperd1 = Decimal(data['netperd1'])
                        budget_line.netperd2 = Decimal(data['netperd2'])
                        budget_line.netperd3 = Decimal(data['netperd3'])
                        budget_line.netperd4 = Decimal(data['netperd4'])
                        budget_line.netperd5 = Decimal(data['netperd5'])
                        budget_line.netperd6 = Decimal(data['netperd6'])
                        budget_line.netperd7 = Decimal(data['netperd7'])
                        budget_line.netperd8 = Decimal(data['netperd8'])
                        budget_line.netperd9 = Decimal(data['netperd9'])
                        budget_line.netperd10 = Decimal(data['netperd10'])
                        budget_line.netperd11 = Decimal(data['netperd11'])
                        budget_line.netperd12 = Decimal(data['netperd12'])
                        budget_line.currency = currency_obj
                        budget_line.budget_set = active.budget_set
                        budget_line.last_updated = datetime.now()
                        budget_line.last_updated_by_id = request.user.id
                        budget_line.save()  # Call save() to trigger pre-save signal

                    return redirect('budgets:update', obj_t.id)
            context = {
                'currency': Currency.objects.all(),
                'name': obj_t.account.acctdesc,
                'form': form,
                'obj': line,
                'object': obj_t,
                'assumptions': assumptions,
                'activebs': active.budget_set,
                'department': obj_t.department.name,
                'lines': lines
            }

            return render(request, 'edit-form.html', context)
    else:
        return redirect('budgets:login')


def change_year(request, year):
    FinancialYear.objects.filter(year=year).update(is_active=True)
    FinancialYear.objects.exclude(year=year).update(is_active=False)
    return redirect("budgets:settings")


def delete_assumption(request, id):
    BudgetAssumptions.objects.filter(id=id).delete()
    return redirect('budgets:assumptions')


def delete_currency(request, id):
    Currency.objects.filter(id=id).delete()
    return redirect('budgets:assumptions')


def update_expenses(request, department_id):
    if request.user.is_authenticated:
        active = get_object_or_404(BudgetVariations, is_active=True)
        department = get_object_or_404(Department, id=department_id)
        year = FinancialYear.objects.get(is_active=True)
        if active:
            if request.user.role != '002':
                single_object = BudgetTotals.objects.filter(budget_set=active.budget_set,
                                                            department=request.user.department, year=year.year)
            else:
                single_object = BudgetTotals.objects.filter(budget_set=active.budget_set, year=year.year)

            first_dept_obj = BudgetTotals.objects.filter(budget_set=active.budget_set,
                                                         department_id=department_id, year=year.year).first()
            user = get_object_or_404(Users, id=request.user.id)
            obj = get_object_or_404(BudgetTotals, id=first_dept_obj.id)
            if request.user.role != '002':
                assumptions = BudgetAssumptions.objects.filter(department=request.user.department)
            else:
                assumptions = BudgetAssumptions.objects.all()
            next_obj = BudgetTotals.objects.filter(id__gt=obj.id, department_id=department_id,
                                                   budget_set=active.budget_set, year=year.year).order_by('id').first()
            prev_obj = BudgetTotals.objects.filter(id__lt=obj.id, department_id=department_id,
                                                   budget_set=active.budget_set, year=year.year).order_by('-id').first()

            first_obj = BudgetTotals.objects.filter(department_id=department_id,
                                                    budget_set=active.budget_set, year=year.year).first()
            last_obj = BudgetTotals.objects.filter(department_id=department_id, budget_set=active.budget_set,
                                                   year=year.year).last()

            form = BudgetEditForm()
            if request.method == 'POST':

                data = json.loads(request.body)
                currency_obj = get_object_or_404(Currency, id=data['currency'])
                with transaction.atomic():
                    BudgetTotals.objects.filter(id=obj.id).update(last_updated=datetime.now(),
                                                                  last_updated_by_id=request.user.id,
                                                                  year=year.year,
                                                                  total=F('total') + data['total'],
                                                                  netperd1=F('netperd1') + data['netperd1'],
                                                                  netperd2=F('netperd2') + data['netperd2'],
                                                                  netperd3=F('netperd3') + data['netperd3'],
                                                                  netperd4=F('netperd4') + data['netperd4'],
                                                                  netperd5=F('netperd5') + data['netperd5'],
                                                                  netperd6=F('netperd6') + data['netperd6'],
                                                                  netperd7=F('netperd7') + data['netperd7'],
                                                                  netperd8=F('netperd8') + data['netperd8'],
                                                                  netperd9=F('netperd9') + data['netperd9'],
                                                                  netperd10=F('netperd10') + data['netperd10'],
                                                                  netperd11=F('netperd11') + data['netperd11'],
                                                                  netperd12=F('netperd12') + data['netperd12']
                                                                  )
                if data['entryType'] == 'function':

                    BudgetLines.objects.create(
                        account=obj.account,
                        year=year.year,
                        exchange_rate=data['exchange_rate'],
                        item_description=data['item_description'],
                        department=obj.department,
                        total=Decimal(data['total']),
                        rate=Decimal(data['rate']),
                        usage=Decimal(data['usage']),
                        factor=Decimal(data['factor']),
                        staff=Decimal(data['staff']),
                        netperd1=Decimal(data['netperd1']),
                        netperd2=Decimal(data['netperd2']),
                        netperd3=Decimal(data['netperd3']),
                        netperd4=Decimal(data['netperd4']),
                        netperd5=Decimal(data['netperd5']),
                        netperd6=Decimal(data['netperd6']),
                        netperd7=Decimal(data['netperd7']),
                        netperd8=Decimal(data['netperd8']),
                        netperd9=Decimal(data['netperd9']),
                        netperd10=Decimal(data['netperd10']),
                        netperd11=Decimal(data['netperd11']),
                        netperd12=Decimal(data['netperd12']),
                        currency=currency_obj,
                        budget_set=active.budget_set,
                        last_updated=datetime.now(),
                        last_updated_by_id=request.user.id)
                    return JsonResponse({'status': 'success'})
                elif data['entryType'] == 'manual':

                    BudgetLines.objects.create(
                        account=obj.account,
                        year=year.year,
                        exchange_rate=data['exchange_rate'],
                        department=obj.department,
                        total=Decimal(data['total']),
                        item_description=data['item_description'],
                        rate=Decimal(0),
                        usage=Decimal(0),
                        factor=Decimal(0),
                        staff=Decimal(0),
                        netperd1=Decimal(data['netperd1']),
                        netperd2=Decimal(data['netperd2']),
                        netperd3=Decimal(data['netperd3']),
                        netperd4=Decimal(data['netperd4']),
                        netperd5=Decimal(data['netperd5']),
                        netperd6=Decimal(data['netperd6']),
                        netperd7=Decimal(data['netperd7']),
                        netperd8=Decimal(data['netperd8']),
                        netperd9=Decimal(data['netperd9']),
                        netperd10=Decimal(data['netperd10']),
                        netperd11=Decimal(data['netperd11']),
                        netperd12=Decimal(data['netperd12']),
                        currency=currency_obj,
                        budget_set=active.budget_set,
                        last_updated=datetime.now(),
                        last_updated_by_id=request.user.id)

                    return JsonResponse({'status': 'success'})
            context = {
                'form': form,
                'obj': single_object,
                'name': obj.account.acctdesc,
                'object': obj,
                'next_obj': next_obj,
                'prev_obj': prev_obj,
                'first_obj': first_obj,
                'last_obj': last_obj,
                'lines': BudgetLines.objects.filter(account__acctid=obj.account.acctid),
                'currency': Currency.objects.all(),
                'assumptions': assumptions,
                'activebs': active.budget_set,
                'department': department.name

            }

            return render(request, 'forms.html', context)
        return render(request, 'forms.html')
    else:
        return redirect('budgets:login')


def saveComments(request):
    if request.method == 'GET':
        budget = request.POST.get('budget_id')
        field_name = request.POST.get('field_name')
        new_comment = request.POST.get('comment')
        obj = get_object_or_404(BudgetComments, budget=budget, field_name=field_name)
        obj.comment = new_comment
        obj.save()
        return JsonResponse({'status': 'success', 'data': new_comment})
    else:
        return JsonResponse({'error': 'Unsupported method'}, status=405)

def currentDate():
    # Get the current date
    current_date = datetime.now()

    # Format the date as YYYYMMDD
    formatted_date = current_date.strftime('%Y%m%d')

    # Convert to BigDecimal
    bigdecimal_date = Decimal(formatted_date)

    return bigdecimal_date

def currentTime():
    """
    Returns the current time in milliseconds since the epoch (January 1, 1970).
    """
    return int(time.time() * 1000)

def post_to_sage(request, budget_set):
    year = FinancialYear.objects.get(is_active=True)
    objects = BudgetTotals.objects.filter(budget_set=budget_set, year=year.year).exclude(total=0)
    print(objects)
    bdgt_set = budget_set
    bs = '1'
    if bdgt_set == 'Budget 2':
        bs = '2'
    elif bdgt_set == 'Budget 3':
        bs = '3'
    for obj in objects:
        net_netperds = [getattr(obj, f'netperd{i}') for i in range(1, 13)]

        # Determine if the object should be saved as positive or negative
        try:
            glafs_obj = Glafs.objects.using('sql_server').get(
                acctid=obj.account_id,
                fscsdsg='1',
                fscscurn='ZMW',
                curntype='F',
                fscsyr=year.year
            )
        except Glafs.DoesNotExist:
            # Handle the case where the Glafs object does not exist
            Glafs.objects.using('sql_server').create(
                acctid=obj.account_id,
                fscsyr=year.year,
                fscsdsg='1',
                fscscurn='ZMW',
                curntype='F',
                audtdate=currentDate(),
                audttime=datetime.now().strftime('%H%M%S'),
                audtuser="ADMIN",
                audtorg="INFDAT",
                swrvl=0,
                codervl="",
                scurndec="2",
                openbal=0.0,
                netperd1=net_netperds[0],
                netperd2=net_netperds[1],
                netperd3=net_netperds[2],
                netperd4=net_netperds[3],
                netperd5=net_netperds[4],
                netperd6=net_netperds[5],
                netperd7=net_netperds[6],
                netperd8=net_netperds[7],
                netperd9=net_netperds[8],
                netperd10=net_netperds[9],
                netperd11=net_netperds[10],
                netperd12=net_netperds[11],
                netperd13=0.0,
                netperd14=0.0,
                netperd15=0.0,
                activitysw=1


            )
            # Compare the netperds values

        glafs_netperds = [getattr(glafs_obj, f'netperd{i}') for i in range(1, 13)]
        if net_netperds != glafs_netperds:
            # For other objects, save them as negative
            total = obj.total
            print(obj)
            sql_obj = Glafs.objects.using('sql_server').filter(acctid=obj.account_id, fscsdsg='1', fscscurn='ZMW',
                                                               curntype='F',
                                                               fscsyr=year.year).update(

                audtdate=datetime.now().date().strftime('%Y%m%d'),
                audttime=datetime.now().strftime('%H%M%S'),
                netperd1=net_netperds[0],
                netperd2=net_netperds[1],
                netperd3=net_netperds[2],
                netperd4=net_netperds[3],
                netperd5=net_netperds[4],
                netperd6=net_netperds[5],
                netperd7=net_netperds[6],
                netperd8=net_netperds[7],
                netperd9=net_netperds[8],
                netperd10=net_netperds[9],
                netperd11=net_netperds[10],
                netperd12=net_netperds[11],

            )

    return redirect('budgets:dashboard-home', )
def post_to_sage_single(request, id):
    year = FinancialYear.objects.get(is_active=True)
    object = BudgetTotals.objects.get(id=id, year=year.year).exclude(total=0)
    sql_obj = Glafs.objects.using('sql_server').filter(acctid=object.account_id, fscsdsg='1', fscscurn='ZMW',
                                                       curntype='F',
                                                       fscsyr=year.year).update(

        audtdate=datetime.now().date().strftime('%Y%m%d'),
        audttime=datetime.now().strftime('%H%M%S'),
        netperd1=object.netperd1,
        netperd2=object.netperd2,
        netperd3=object.netperd3,
        netperd4=object.netperd4,
        netperd5=object.netperd5,
        netperd6=object.netperd6,
        netperd7=object.netperd7,
        netperd8=object.netperd8,
        netperd9=object.netperd9,
        netperd10=object.netperd10,
        netperd11=object.netperd11,
        netperd12=object.netperd12,

    )

    return redirect('budgets:update',id )


"""
def year_aggregate(queryset, queryset2, queryset3, queryset4, dept_id):
    period_values_dict = {}
    for current_q in range(1, 5):  # for each quarter
        quarter_values = total_quarter_aggregate(queryset, queryset2, queryset3, queryset4, current_q, dept_id)
        period_values_dict[f'Q{current_q}'] = quarter_values
    for current_h in range(1, 3):  # for each half year
        half_values = half_year_aggregate(queryset, queryset2, queryset3, queryset4, current_h, dept_id)
        period_values_dict[f'H{current_h}'] = half_values
    year_values = year_total_aggregate(queryset, queryset2, queryset3, queryset4, dept_id)
    period_values_dict['YRT'] = year_values
    return period_values_dict


def half_year_aggregate(queryset, queryset2, queryset3, queryset4, current_q, dept_id, ):
    period_values_dict = {
        'total': Decimal('0.00'),
        'total_actual': Decimal('0.00'),
        'available': Decimal('0.00'),
        'pending': Decimal('0.00'),
        'year_total': Decimal('0.00'),
        'year_total_actual': Decimal('0.00'),
        'pending_year_total': Decimal('0.00'),
        'available_year_total': Decimal('0.00')
    }

    ranges = {
        1: 'audtdate__gte=20240101,audtdate__lte=20240631',

        2: 'audtdate__gte=20240701,audtdate__lte=20241231',

    }
    months_in_current_half = netperd_in_half(current_q)
    net_perds = [month for month in months_in_current_half]
    period_values_aggregations = {period: Sum(period) for period in net_perds}

    if queryset and queryset2:
        period_values_sum = queryset.aggregate(**period_values_aggregations)
        netperd_fields = [f"netperd{i}" for i in range(1, 13)]
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = queryset2.aggregate(**total_period_values_aggregations)
        quarter_total = sum(period_values_sum.values())
        year_total = sum(total_period_values_sum.values())
        start_date, end_date = ranges[int(current_q)].split(',')
        start_date = int(start_date.split('=')[1])
        end_date = int(end_date.split('=')[1])
        acctids = [obj.account_id for obj in queryset]
        pending_expenses = Enebd.objects.using('sql_server').filter(
            idglacct__in=acctids,
            status=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_expenses_all = Enebd.objects.using('sql_server').filter(
            idglacct__in=acctids,
            status=1
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_requisition = Enrqnl.objects.using('sql_server').filter(
            glacct__in=acctids,
            status=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_requisition_all = Enrqnl.objects.using('sql_server').filter(
            glacct__in=acctids,
            status=1
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order = Poporl.objects.using('sql_server').filter(
            glnonstkcr__in=acctids,
            completion=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order_all = Poporl.objects.using('sql_server').filter(
            glnonstkcr__in=acctids,
            completion=1
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_total = pending_expenses + pending_requisition + pending_purchase_order
        pending_year_total = pending_expenses_all + pending_requisition_all + pending_purchase_order_all
        period_values_dict['pending'] = round(pending_total, 2)
        period_values_dict['pending_year_total'] = round(pending_year_total, 2)
        period_values_dict['available'] = round(quarter_total, 2) - round(pending_total, 2)
        period_values_dict['available_year_total'] = round(year_total, 2) - round(pending_year_total, 2)
        period_values_dict['total'] = round(quarter_total, 2)
        period_values_dict['year_total'] = round(year_total, 2)

    if queryset3 and queryset4:
        period_values_sum = queryset3.aggregate(**period_values_aggregations)
        netperd_fields = [f"netperd{i}" for i in range(1, 13)]
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = queryset4.aggregate(
            **total_period_values_aggregations)
        quarter_total_actual = sum(period_values_sum.values())
        year_total_actual = sum(total_period_values_sum.values())

        period_values_dict['total_actual'] = round(quarter_total_actual, 2)
        period_values_dict['year_total_actual'] = round(year_total_actual, 2)

    return period_values_dict


def year_total_aggregate(queryset, queryset2, queryset3, queryset4, dept_id, ):
    period_values_dict = {
        'total': Decimal('0.00'),
        'total_actual': Decimal('0.00'),
        'available': Decimal('0.00'),
        'pending': Decimal('0.00'),
        'year_total': Decimal('0.00'),
        'year_total_actual': Decimal('0.00'),
        'pending_year_total': Decimal('0.00'),
        'available_year_total': Decimal('0.00')
    }

    ranges = {
        1: ',',

    }
    netperd_fields = [f"netperd{i}" for i in range(1, 13)]
    total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
    if queryset and queryset2:
        total_period_values_sum = queryset2.aggregate(**total_period_values_aggregations)
        period_values_sum = queryset.aggregate(**total_period_values_aggregations)
        quarter_total = sum(period_values_sum.values())
        year_total = sum(total_period_values_sum.values())
        start_date = int(20240101)
        end_date = int(20241231)
        acctids = [obj.account_id for obj in queryset]
        pending_expenses = Enebd.objects.using('sql_server').filter(
            idglacct__in=acctids,
            status=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_expenses_all = Enebd.objects.using('sql_server').filter(
            idglacct__in=acctids,
            status=1
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_requisition = Enrqnl.objects.using('sql_server').filter(
            glacct__in=acctids,
            status=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_requisition_all = Enrqnl.objects.using('sql_server').filter(
            glacct__in=acctids,
            status=1
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order = Poporl.objects.using('sql_server').filter(
            glnonstkcr__in=acctids,
            completion=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order_all = Poporl.objects.using('sql_server').filter(
            glnonstkcr__in=acctids,
            completion=1
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_total = pending_expenses + pending_requisition + pending_purchase_order
        pending_year_total = pending_expenses_all + pending_requisition_all + pending_purchase_order_all
        period_values_dict['pending'] = round(pending_total, 2)
        period_values_dict['pending_year_total'] = round(pending_year_total, 2)
        period_values_dict['available'] = round(quarter_total, 2) - round(pending_total, 2)
        period_values_dict['available_year_total'] = round(year_total, 2) - round(pending_year_total, 2)
        period_values_dict['total'] = round(quarter_total, 2)
        period_values_dict['year_total'] = round(year_total, 2)

    if queryset3 and queryset4:
        period_values_sum = queryset3.aggregate(**total_period_values_aggregations)
        netperd_fields = [f"netperd{i}" for i in range(1, 13)]
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = queryset4.aggregate(
            **total_period_values_aggregations)
        quarter_total_actual = sum(period_values_sum.values())
        year_total_actual = sum(total_period_values_sum.values())

        period_values_dict['total_actual'] = round(quarter_total_actual, 2)
        period_values_dict['year_total_actual'] = round(year_total_actual, 2)

    return period_values_dict


def total_quarter_aggregate(queryset, queryset2, queryset3, queryset4, current_q, dept_id, ):
    period_values_dict = {
        'total': Decimal('0.00'),
        'total_actual': Decimal('0.00'),
        'available': Decimal('0.00'),
        'pending': Decimal('0.00'),
        'year_total': Decimal('0.00'),
        'year_total_actual': Decimal('0.00'),
        'pending_year_total': Decimal('0.00'),
        'available_year_total': Decimal('0.00')
    }

    ranges = {
        1: 'audtdate__gte=20240101,audtdate__lte=20240331',

        2: 'audtdate__gte=20240401,audtdate__lte=20240630',

        3: 'audtdate__gte=20240701,audtdate__lte=20240930',

        4: 'audtdate__gte=20241001,audtdate__lte=20241231',

    }
    months_in_current_quarter = netperd_in_quarter(current_q)
    net_perds = [month for month in months_in_current_quarter]
    period_values_aggregations = {period: Sum(period) for period in net_perds}

    if queryset and queryset2:
        period_values_sum = queryset.aggregate(**period_values_aggregations)
        netperd_fields = [f"netperd{i}" for i in range(1, 13)]
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = queryset2.aggregate(**total_period_values_aggregations)
        quarter_total = sum(period_values_sum.values())
        year_total = sum(total_period_values_sum.values())
        start_date, end_date = ranges[int(current_q)].split(',')
        start_date = int(start_date.split('=')[1])
        end_date = int(end_date.split('=')[1])
        acctids = [obj.account_id for obj in queryset]
        pending_expenses = Enebd.objects.using('sql_server').filter(
            idglacct__in=acctids,
            status=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_expenses_all = Enebd.objects.using('sql_server').filter(
            idglacct__in=acctids,
            status=1
        ).aggregate(total_amount=Sum('amtlinet'))['total_amount'] or 0

        pending_requisition = Enrqnl.objects.using('sql_server').filter(
            glacct__in=acctids,
            status=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_requisition_all = Enrqnl.objects.using('sql_server').filter(
            glacct__in=acctids,
            status=1
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order = Poporl.objects.using('sql_server').filter(
            glnonstkcr__in=acctids,
            completion=1,
            audtdate__gte=start_date,
            audtdate__lte=end_date
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_purchase_order_all = Poporl.objects.using('sql_server').filter(
            glnonstkcr__in=acctids,
            completion=1
        ).aggregate(total_amount=Sum('extended'))['total_amount'] or 0

        pending_total = pending_expenses + pending_requisition + pending_purchase_order
        pending_year_total = pending_expenses_all + pending_requisition_all + pending_purchase_order_all
        period_values_dict['pending'] = round(pending_total, 2)
        period_values_dict['pending_year_total'] = round(pending_year_total, 2)
        period_values_dict['available'] = round(quarter_total, 2) - round(pending_total, 2)
        period_values_dict['available_year_total'] = round(year_total, 2) - round(pending_year_total, 2)
        period_values_dict['total'] = round(quarter_total, 2)
        period_values_dict['year_total'] = round(year_total, 2)

    if queryset3 and queryset4:
        period_values_sum = queryset3.aggregate(**period_values_aggregations)
        netperd_fields = [f"netperd{i}" for i in range(1, 13)]
        total_period_values_aggregations = {period: Sum(period) for period in netperd_fields}
        total_period_values_sum = queryset4.aggregate(
            **total_period_values_aggregations)
        quarter_total_actual = sum(period_values_sum.values())
        year_total_actual = sum(total_period_values_sum.values())

        period_values_dict['total_actual'] = round(quarter_total_actual, 2)
        period_values_dict['year_total_actual'] = round(year_total_actual, 2)

    return period_values_dict
"""
